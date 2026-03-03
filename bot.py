"""
Payment Tracker Bot v3
- Full context memory (context.txt)
- Confirmation before writing to Excel
- Balance reconciliation
- Context view/edit via Telegram
"""
import os, json, logging, base64
from datetime import datetime, time, date as date_type, timedelta
from pathlib import Path
import httpx
from openpyxl import load_workbook
import io
try:
    from pypdf import PdfReader
    HAS_PDF = True
except ImportError:
    HAS_PDF = False
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from telegram import Update, Bot, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (Application, MessageHandler, CommandHandler,
                           CallbackQueryHandler, filters, ContextTypes)

BOT_TOKEN     = os.environ["BOT_TOKEN"]
ANTHROPIC_KEY = os.environ["ANTHROPIC_KEY"]
MY_CHAT_ID    = int(os.environ["MY_CHAT_ID"])
MORNING_HOUR  = int(os.environ.get("MORNING_HOUR", "9"))

DATA_FILE    = Path("data/messages.json")
EXCEL_FILE   = Path("Agent_Model_v2.xlsx")
CONTEXT_FILE = Path("data/context.txt")
PENDING_FILE = Path("data/pending_update.json")  # stores parsed data awaiting confirmation

logging.basicConfig(level=logging.INFO)

def _ensure_settings_usdt():
    """Add USDT=1.0 to Settings FX table if not present."""
    if not EXCEL_FILE.exists(): return
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Settings"]
        # Check if USDT already there
        for row in ws.iter_rows(min_row=7, max_row=25, values_only=True):
            if row[0] == "USDT": return
        # Add after last FX entry
        for r in range(7, 25):
            if not ws.cell(r, 1).value:
                ws.cell(r, 1).value = "USDT"
                ws.cell(r, 2).value = 1.0
                from openpyxl.styles import Font, Alignment
                ws.cell(r, 1).font = Font(name="Arial", size=9)
                ws.cell(r, 2).font = Font(name="Arial", size=9)
                ws.cell(r, 2).alignment = Alignment(horizontal="right")
                wb.save(EXCEL_FILE)
                log.info("Added USDT=1.0 to Settings FX table")
                return
    except Exception as e:
        log.error(f"_ensure_settings_usdt: {e}")

log = logging.getLogger(__name__)

# ── Styles ────────────────────────────────────────────────────────────────────
WHITE  = "FFFFFF"; YELLOW = "FFF2CC"; GREEN  = "E2EFDA"
RED    = "FCE4D6"; ORANGE = "FDEBD0"; LIGHT  = "D6E4F0"; LGRAY  = "F2F2F2"
BLUE_LIGHT = "BDD7EE"

# ── Agent company filter ──────────────────────────────────────────────────────
# Fuzzy substring tokens — "Balkemy GT", "BALKEMY GENERAL TRADING", "from Balkemy"
# all match because "balkemy" is a substring.
_AGENT_CO_TOKENS = frozenset({
    "balkemy", "troveco", "elitesphere", "rawrima", "masadan", "gornik",
    "nexus marine", "asteno",
})

def is_agent_company_str(val) -> bool:
    """Returns True if val looks like one of the agent's companies.
    Uses fuzzy substring matching so "Balkemy GT" and "BALKEMY GENERAL TRADING"
    both return True. Agent companies should never be written as beneficiary."""
    if not val: return False
    lo = str(val).lower()
    return any(token in lo for token in _AGENT_CO_TOKENS)
thin   = Side(style="thin", color="BFBFBF")
def B(): return Border(top=thin, bottom=thin, left=thin, right=thin)
TYPE_BG = {"Deposit": GREEN, "Payment": WHITE, "Cash Out": ORANGE,
           "Cash In": LIGHT, "❓ Unknown": RED}
STAT_BG = {"✅ Paid": GREEN, "⏳ Pending": YELLOW,
           "🔄 In Progress": BLUE_LIGHT,
           "⚠ Partial/Check": ORANGE, "❓ Clarify": RED}

def sc(cell, bg=WHITE, bold=False, sz=9, fc="000000", num=None,
       align="left", wrap=False):
    cell.font      = Font(name="Arial", bold=bold, size=sz, color=fc)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border    = B()
    if num: cell.number_format = num

# ── Context ───────────────────────────────────────────────────────────────────
def load_context() -> str:
    CONTEXT_FILE.parent.mkdir(exist_ok=True)
    return CONTEXT_FILE.read_text(encoding="utf-8") if CONTEXT_FILE.exists() else ""

def save_context(text: str):
    CONTEXT_FILE.parent.mkdir(exist_ok=True)
    CONTEXT_FILE.write_text(text, encoding="utf-8")

def update_context_after_update(new_info: str):
    """Append new findings to context file."""
    ctx = load_context()
    ts  = datetime.now().strftime("%d.%m.%Y %H:%M")
    ctx += f"\n\n--- ОБНОВЛЕНИЕ {ts} ---\n{new_info}"
    save_context(ctx)


CHAT_HISTORY_FILE = Path("data/chat_history.json")
MAX_HISTORY = 8  # keep last 8 exchanges

def load_chat_history():
    CHAT_HISTORY_FILE.parent.mkdir(exist_ok=True)
    return json.loads(CHAT_HISTORY_FILE.read_text(encoding="utf-8")) if CHAT_HISTORY_FILE.exists() else []

def save_chat_history(history: list):
    # Keep only last MAX_HISTORY exchanges
    if len(history) > MAX_HISTORY * 2:
        history = history[-(MAX_HISTORY * 2):]
    CHAT_HISTORY_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")

def clear_chat_history():
    if CHAT_HISTORY_FILE.exists(): CHAT_HISTORY_FILE.unlink()

def get_recent_transactions_with_rows(n=10) -> str:
    """Get last N transactions with row numbers for chat editing."""
    if not EXCEL_FILE.exists(): return ""
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["Transactions"]
        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=5, max_col=12, values_only=True), start=5):
            if row[0]: rows.append((i, row))
        lines = []
        for i, row in rows[-n:]:
            lines.append(f"row={i} | {row[0]} | {row[1]} | {str(row[2] or '')[:40]} | {row[4]} {row[5]} | K={row[10]}")
        return "\n".join(lines)
    except: return ""

def get_excel_summary() -> str:
    """Get compact Excel state for chat context."""
    lines = []
    bal = get_balance_from_excel()
    if bal:
        lines.append(f"Баланс агента: ${bal[0]:,.2f} USD (на {bal[1]})")
    pending, usd_total, tbc_count = get_pending_invoices()
    if pending:
        lines.append(f"Pending инвойсов: {len(pending)}")
        lines.extend(pending[:5])
        if len(pending) > 5:
            lines.append(f"  ...и ещё {len(pending)-5}")
        tbc_note = f" + {tbc_count} с суммой TBC" if tbc_count else ""
        lines.append(f"  Итого к оплате: ~${usd_total:,.0f} USD{tbc_note}")
    unknown = get_unknown_transactions()
    if unknown:
        lines.append(f"Неизвестных транзакций: {len(unknown)}")
    queue = load_messages()
    if queue:
        lines.append(f"Накоплено сообщений от агента: {len(queue)} (для /update)")
    return "\n".join(lines) if lines else "Excel не найден"

# ── Message store ─────────────────────────────────────────────────────────────
def load_messages():
    DATA_FILE.parent.mkdir(exist_ok=True)
    return json.loads(DATA_FILE.read_text(encoding="utf-8")) if DATA_FILE.exists() else []

def save_message(d):
    msgs = load_messages(); msgs.append(d)
    DATA_FILE.write_text(json.dumps(msgs, ensure_ascii=False, indent=2), encoding="utf-8")

def clear_messages():
    DATA_FILE.write_text("[]", encoding="utf-8")

def _fmt(msgs):
    lines = []
    for m in msgs:
        line = f"[{m['date']}] {m.get('sender','?')}:"
        if m.get("text"): line += f" {m['text']}"
        if m.get("file"): line += f" [файл: {m['file']}]"
        if m.get("pdf_content"):
            line += f"\n  [СОДЕРЖИМОЕ PDF {m['file']}]:\n  {m['pdf_content'][:2000]}"
        lines.append(line)
    return "\n".join(lines)

def _clean_json(raw: str) -> str:
    """Strip markdown fences and leading 'json' tag from Claude's response."""
    raw = raw.strip().strip("```").strip()
    if raw.startswith("json"): raw = raw[4:].strip()
    return raw

def _build_multimodal_content(msgs: list) -> list:
    """
    Build a multimodal content list for Claude API.
    Text messages → text blocks.
    PDFs with pdf_b64 → document blocks (native Claude PDF reading).
    Images with img_b64 → image blocks (native Claude vision).
    PDFs with only extracted text → included in text block as fallback.
    Returns list of content blocks ready for the 'content' field.
    """
    content = []
    text_parts = []

    for m in msgs:
        line = f"[{m['date']}] {m.get('sender', '?')}:"
        if m.get("text"):
            line += f" {m['text']}"
        if m.get("file") and not m.get("pdf_b64") and not m.get("img_b64"):
            # No b64 — include extracted text as fallback
            line += f" [файл: {m['file']}]"
            if m.get("pdf_content"):
                line += f"\n  [ТЕКСТ PDF]:\n  {m['pdf_content'][:2000]}"
        elif m.get("pdf_b64"):
            line += f" [PDF: {m['file']} — содержимое ниже как документ]"
        elif m.get("img_b64"):
            line += f" [Изображение: {m['file']} — содержимое ниже]"
        text_parts.append(line)

    # Text block first (required — context before documents/images)
    if text_parts:
        content.append({"type": "text", "text": "\n".join(text_parts)})

    # Each PDF as a native document block
    for m in msgs:
        if m.get("pdf_b64"):
            content.append({
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": m["pdf_b64"],
                },
                "title": m.get("file", "invoice.pdf"),
                "cache_control": {"type": "ephemeral"},
            })

    # Each image as a native image block
    for m in msgs:
        if m.get("img_b64"):
            content.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": m.get("img_media", "image/jpeg"),
                    "data": m["img_b64"],
                },
            })

    return content

# ── Pending confirmation store ────────────────────────────────────────────────
def save_pending(data: dict):
    PENDING_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def load_pending() -> dict:
    return json.loads(PENDING_FILE.read_text(encoding="utf-8")) if PENDING_FILE.exists() else {}

def clear_pending():
    if PENDING_FILE.exists(): PENDING_FILE.unlink()

# ── Excel read ────────────────────────────────────────────────────────────────
def get_balance_from_excel():
    if not EXCEL_FILE.exists(): return None
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["Transactions"]
        last_bal = last_date = None
        for row in ws.iter_rows(min_row=5, max_col=11, values_only=True):
            if row[10] is not None and isinstance(row[10], (int, float)):
                last_bal  = row[10]
                last_date = row[0]
        return (last_bal, last_date) if last_bal else None
    except Exception as e:
        log.error(f"Excel balance: {e}"); return None

def _compute_usd(wb, ccy, amount):
    """Compute USD equivalent from FX Settings — no formula cache needed."""
    if not isinstance(amount, (int, float)) or amount <= 0:
        return None
    fx = 1.0
    try:
        for row in wb["Settings"].iter_rows(min_row=7, max_row=25, values_only=True):
            if row[0] and str(row[0]) == str(ccy) and isinstance(row[1], (int,float)):
                fx = float(row[1]); break
    except Exception: pass
    return round(amount / fx, 2) if fx else amount


def get_pending_invoices():
    """Returns (lines, usd_total, tbc_count) for all non-paid invoices.
    Computes USD from Settings FX table — works even after openpyxl save clears formula cache."""
    if not EXCEL_FILE.exists(): return [], 0.0, 0
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["Invoices"]
        out = []
        usd_total = 0.0
        tbc_count = 0
        for row in ws.iter_rows(min_row=5, max_col=11, values_only=True):
            if row[6] and row[6] != "✅ Paid" and (row[0] or row[1]):
                amt_raw = row[4]
                ccy     = str(row[3] or "")
                amt     = f"{amt_raw:,.2f}" if isinstance(amt_raw, (int,float)) else str(amt_raw or "TBC")
                # Try F col first (may have computed number); fall back to Python calc
                usd_val = row[5] if isinstance(row[5], (int,float)) else None
                if usd_val is None and isinstance(amt_raw, (int,float)):
                    usd_val = _compute_usd(wb, ccy, amt_raw)
                if isinstance(usd_val, (int, float)) and usd_val > 0:
                    usd_str = f" ≈ ${usd_val:,.0f}"
                    usd_total += usd_val
                else:
                    usd_str = " (USD TBC)"
                    tbc_count += 1
                benef_str = f" | for: {row[10]}" if row[10] else ""
                out.append(f"- {row[2] or '?'}: {amt} {ccy}{usd_str}{benef_str}")
        return out, usd_total, tbc_count
    except Exception as e:
        log.error(f"Excel pending: {e}"); return [], 0.0, 0


def get_recent_unconfirmed(days=14):
    """Get recent Cash In / Deposits that might not be confirmed by agent yet."""
    if not EXCEL_FILE.exists(): return []
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["Transactions"]
        items = []
        for row in ws.iter_rows(min_row=5, max_col=12, values_only=True):
            if row[1] in ("Cash In", "Deposit") and row[0]:
                note = str(row[11] or "").lower()
                # Flag as potentially unconfirmed if notes say so or if recent
                if "unconfirmed" in note or "follow up" in note or "not confirm" in note:
                    amt = f"{row[5]:,.2f}" if isinstance(row[5],(int,float)) else str(row[5])
                    items.append(f"- {row[0]}: {row[2] or ''} | {amt} {row[4] or ''} | ПРИМЕЧАНИЕ: {row[11]}")
        return items
    except Exception as e:
        log.error(f"get_recent_unconfirmed error: {e}"); return []


def get_existing_invoices_list():
    """Return list of existing invoice IDs and payees for dedup check.
    Includes Excel row numbers so Claude can target edits correctly."""
    if not EXCEL_FILE.exists(): return ""
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        wi = wb["Invoices"]
        lines = []
        for i, row in enumerate(wi.iter_rows(min_row=5, values_only=True), start=5):
            if row[1] or row[2]:
                benef_str = f" | for={row[10]}" if len(row) > 10 and row[10] else ""
                lines.append(f"row={i} | inv={row[1] or '?'} | payee={row[2] or '?'} | ccy={row[3]} | amt={row[4]} | status={row[6]}{benef_str}")
        return "\n".join(lines)
    except Exception as e:
        log.error(f"get_existing_invoices: {e}"); return ""

def get_recent_transactions(days: int = 60) -> list:
    """Read recent transactions from Excel for dedup check."""
    if not EXCEL_FILE.exists():
        return []
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["Transactions"]
        cutoff = date_type.today() - timedelta(days=days)
        result = []
        for row in ws.iter_rows(min_row=5, max_col=8, values_only=True):
            if not row[0]:
                continue
            d = _parse_date(row[0])
            if d and d >= cutoff:
                amt = row[5]  # col F: original amount
                ccy = str(row[4] or "")
                if isinstance(amt, (int, float)) and amt > 0 and ccy:
                    result.append({"date": d, "ccy": ccy, "amount": float(amt)})
        return result
    except Exception as e:
        log.error(f"get_recent_transactions: {e}")
        return []


def _is_duplicate_tx(tx_candidate: dict, existing_txs: list) -> tuple:
    """
    Check if a transaction from Claude's output already exists in Excel.
    Match criteria: same currency + amount within 1% + date within window.
    Window: 60 days for Deposit/Cash In (agent confirms later than receipt),
            7 days for other types.
    """
    ccy = str(tx_candidate.get("ccy", "")).upper()
    try:
        amt = float(tx_candidate.get("amount", 0))
    except (TypeError, ValueError):
        return False, ""
    if amt <= 0 or not ccy:
        return False, ""

    date_str = tx_candidate.get("date", "")
    tx_date = _parse_date(date_str) if date_str else None
    tx_type = str(tx_candidate.get("type", "")).lower()
    # Deposits and Cash In can appear in agent statements weeks after recording
    date_window = 60 if any(k in tx_type for k in ("deposit", "cash in", "receipt")) else 7

    for ex in existing_txs:
        if ex["ccy"].upper() != ccy:
            continue
        if abs(ex["amount"] - amt) / max(ex["amount"], 1) > 0.01:
            continue
        if tx_date and ex["date"]:
            if abs((tx_date - ex["date"]).days) > date_window:
                continue
        return True, f"{ccy} {amt:,.2f} уже в Excel ({ex['date']})"

    return False, ""


def _dedup_transactions(data: dict) -> tuple:
    """
    Remove duplicate transactions from Claude's output.
    Returns (cleaned_data, skipped_descriptions).
    """
    txs = data.get("new_transactions", [])
    if not txs:
        return data, []

    existing = get_recent_transactions(days=60)
    clean_txs = []
    skipped = []

    for tx in txs:
        is_dup, reason = _is_duplicate_tx(tx, existing)
        if is_dup:
            amt = tx.get("amount", 0)
            ccy = tx.get("ccy", "")
            skipped.append(f"  ↩ {tx.get('date', '')} | {tx.get('type', '')} | "
                           f"{float(amt):,.2f} {ccy} — {reason}")
        else:
            clean_txs.append(tx)

    data = dict(data)
    data["new_transactions"] = clean_txs
    return data, skipped


def _get_paid_invoice_nos() -> set:
    """Return set of invoice_no strings that are already ✅ Paid in Excel."""
    if not EXCEL_FILE.exists():
        return set()
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        wi = wb["Invoices"]
        paid = set()
        for row in wi.iter_rows(min_row=5, max_col=10, values_only=True):
            if row[6] == "✅ Paid" and row[1]:
                paid.add(str(row[1]).strip())
        return paid
    except Exception as e:
        log.error(f"_get_paid_invoice_nos: {e}")
        return set()


def _invoice_has_transaction(invoice_no: str) -> bool:
    """Check if a transaction referencing this invoice_no exists in Transactions sheet.
    Builds multiple search variants to handle format differences like
    '2053_RO' stored as 'inv 2053', or 'INV 4410 (5YZ...)' stored as 'INV 4410 — 5YZ...'.
    """
    if not EXCEL_FILE.exists():
        return False
    try:
        import re
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["Transactions"]
        raw = invoice_no.strip()

        variants = set()
        variants.add(raw.lower())

        # PRIMARY: exact inv= tag written at transaction creation time
        variants.add(f"inv={raw.lower()}")

        # Strip _SUFFIX (e.g. 2053_RO → 2053)
        base = raw.split("_")[0].strip()
        if base and base != raw:
            variants.add(base.lower())
            variants.add(f"inv={base.lower()}")

        # Strip everything after first special char/bracket/dash
        # "INV 4410 (5YZ W L L-2)" → "INV 4410"
        short = re.split(r'[\(\[\{—\-–]', raw)[0].strip()
        if short and short != raw:
            variants.add(short.lower())

        # If we have a number in the string, add just that number
        nums = re.findall(r'\d{3,}', raw)
        for n in nums:
            variants.add(n)

        for row in ws.iter_rows(min_row=5, max_col=12, values_only=True):
            for col in (2, 11):
                cell = str(row[col] or "").lower()
                if any(v in cell for v in variants):
                    return True
        return False
    except Exception as e:
        log.error(f"_invoice_has_transaction: {e}")
        return False


def _dedup_invoice_updates(data: dict) -> tuple:
    """
    Remove invoice_updates where invoice is already Paid AND transaction exists.
    If Paid but no transaction — keep with ⚠ warning.
    If In Progress → Paid — always keep (legitimate update).
    Returns (cleaned_data, skipped_descriptions).
    """
    upds = data.get("invoice_updates", [])
    if not upds:
        return data, []

    paid_nos = _get_paid_invoice_nos()
    clean_upds = []
    skipped = []

    for upd in upds:
        inv_no = str(upd.get("invoice_no", "")).strip()
        new_status = upd.get("new_status", "")

        if "Paid" not in new_status:
            clean_upds.append(upd)
            continue

        if inv_no not in paid_nos:
            clean_upds.append(upd)
            continue

        # Already Paid in Excel — check if transaction exists
        has_tx = _invoice_has_transaction(inv_no)
        if has_tx:
            skipped.append(f"  ↩ {inv_no} → уже ✅ Paid в Excel (транзакция есть)")
        else:
            # Keep but warn: manual edit suspected, no transaction
            upd = dict(upd)
            upd["_warning"] = "⚠ Инвойс уже Paid в Excel, но транзакция не найдена — проверь"
            clean_upds.append(upd)

    data = dict(data)
    data["invoice_updates"] = clean_upds
    return data, skipped


def get_unknown_transactions():
    if not EXCEL_FILE.exists(): return []
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["Transactions"]
        out = []
        for row in ws.iter_rows(min_row=5, max_col=12, values_only=True):
            if row[1] and "Unknown" in str(row[1]):
                amt = f"{row[5]:,.2f}" if isinstance(row[5], (int,float)) else str(row[5] or "?")
                out.append(f"- {row[0]}: {row[2] or '?'} | {amt} {row[4] or ''}")
        return out
    except Exception as e:
        log.error(f"Excel unknown: {e}"); return []

# ── Excel write ───────────────────────────────────────────────────────────────
def find_last_row(ws, start=5):
    """Find last data row by checking col A (date) OR col B (type/invoice_no)."""
    last = start - 1
    for row in ws.iter_rows(min_row=start, max_col=2):
        if row[0].value is not None or row[1].value is not None:
            last = row[0].row
    return last

def _get_fx(ws_parent, ccy):
    """Lookup FX rate from Settings sheet."""
    try:
        for row in ws_parent["Settings"].iter_rows(min_row=7, max_row=25, values_only=True):
            if str(row[0]) == ccy and row[1]:
                return float(row[1])
    except Exception: pass
    return 1.0

def _get_comm(tp, ccy):
    """Get commission rate."""
    if ccy == "RUB": return 0.004
    return {"Deposit":0.0,"Cash In":0.0,"Payment":0.005,
            "Cash Out":0.005,"❓ Unknown":0.005}.get(tp, 0.005)

def _prev_balance(ws, r):
    """Get last computed balance before row r."""
    last = 0.0
    for row in ws.iter_rows(min_row=5, max_row=r-1, max_col=11, values_only=True):
        if row[10] is not None and isinstance(row[10], (int,float)):
            last = float(row[10])
    return last

def apply_tx_row(ws, r, tx):
    tp  = tx.get("type", "Payment")
    bg  = TYPE_BG.get(tp, WHITE)
    ccy = tx.get("ccy","USD")

    # Compute values
    try: amt = float(tx.get("amount") or 0)
    except: amt = 0.0

    fx = float(tx.get("fx_rate")) if tx.get("fx_rate") else _get_fx(ws.parent, ccy)
    comm = float(tx.get("comm")) if tx.get("comm") else _get_comm(tp, ccy)
    gross = round(amt / fx, 2) if fx else amt
    net   = round(gross, 2) if tp in ("Deposit","Cash In") else round(-(gross/max(1-comm,0.0001)),2)
    bal   = round(_prev_balance(ws, r) + net, 2)

    # Write base columns A-F, L
    for col_i, val in enumerate([
        tx.get("date",""), tp, tx.get("description",""), tx.get("payee",""),
        ccy, amt, None, None, None, None, None, tx.get("notes","")
    ], 1):
        if col_i in (7,8,9,10,11): continue
        c = ws.cell(r, col_i, val if val is not None else "")
        sc(c, bg=bg, wrap=(col_i in (3,12)), sz=9)

    # G: FX rate
    sc(ws.cell(r, 7, round(fx,5)), bg=YELLOW, fc="0000CC", num="0.00000")
    # H: Gross USD
    sc(ws.cell(r, 8, gross), bg=YELLOW, num="#,##0.00")
    # I: Comm %
    sc(ws.cell(r, 9, comm), bg=YELLOW, fc="0000CC", num="0.0%")
    # J: Net USD — число, не формула (формулы смещаются при copy-paste)
    sc(ws.cell(r, 10, round(net, 2)), bg=YELLOW, num='#,##0.00')
    # K: Balance — число
    sc(ws.cell(r, 11, bal), bg=YELLOW, bold=True, fc="1F3864", num='#,##0.00')
    # M: Payer, N: Beneficiary
    payer = tx.get("payer") or None
    benef = tx.get("beneficiary") or None
    if payer is not None:
        sc(ws.cell(r, 13, payer), bg=bg, sz=9, wrap=False)
    if benef is not None and not is_agent_company_str(benef):
        sc(ws.cell(r, 14, benef), bg=bg, sz=9, wrap=False)
    ws.row_dimensions[r].height = 28

def _parse_date(s):
    """Parse DD.MM.YYYY or similar to date object, return None on fail."""
    from datetime import date as ddate
    if not s: return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            from datetime import datetime
            return datetime.strptime(str(s).strip(), fmt).date()
        except: pass
    return None


def _find_duplicate_tx(wst, payee: str, ccy: str, amount: float, date_str: str,
                       ref: str = None) -> int | None:
    """
    Check if Transactions already has a similar Payment row.
    Returns row number if duplicate found, else None.
    Match criteria: same payee (fuzzy) + same CCY + amount within 1% + date within 10 days.
    EXCLUSION: if ref provided and candidate has a DIFFERENT non-empty ref → not a duplicate
    (two real payments to same payee for same amount with different refs = different invoices).
    """
    from datetime import timedelta
    ref_date = _parse_date(date_str)
    payee_lo = (payee or "").lower().strip()
    ref_lo   = (ref or "").lower().strip()

    for row_idx, row in enumerate(wst.iter_rows(min_row=5, max_col=12, values_only=True), start=5):
        if not row[0]: continue
        r_type = str(row[1] or "")
        if r_type not in ("Payment", "Deposit"): continue

        # Payee match (partial)
        r_payee = str(row[3] or "").lower()
        r_desc  = str(row[2] or "").lower()
        payee_words = [w for w in payee_lo.split() if len(w) > 3]
        payee_match = any(w in r_payee or w in r_desc for w in payee_words) if payee_words else False

        # CCY + amount match (within 1%)
        r_ccy = str(row[4] or "")
        try: r_amt = float(row[5] or 0)
        except: r_amt = 0.0
        amt_match = (r_ccy == ccy and amount > 0 and abs(r_amt - amount) / max(amount, 1) < 0.01)

        # Date proximity (within 10 days)
        date_match = True
        if ref_date:
            r_date = _parse_date(row[0])
            if r_date:
                date_match = abs((r_date - ref_date).days) <= 10

        if payee_match and amt_match and date_match:
            # Ref exclusion: if we have a ref and the candidate has a DIFFERENT ref → skip
            if ref_lo:
                existing_notes = str(row[11] or "").lower()
                existing_ref = ""
                if "ref: " in existing_notes:
                    try:
                        existing_ref = existing_notes.split("ref: ")[1].split("|")[0].strip().split()[0]
                    except: pass
                if existing_ref and existing_ref != ref_lo:
                    continue  # different refs = different transactions
            return row_idx
    return None


def _flag_duplicate(wst, row_a: int, row_b: int):
    """Mark two rows as possible duplicates in their Notes column."""
    ORANGE = PatternFill("solid", fgColor="FFFCE4D6")
    for r in (row_a, row_b):
        if r is None: continue
        c = wst.cell(r, 12)
        cur = str(c.value or "")
        if "POSSIBLE DUPLICATE" not in cur:
            c.value = (cur + " | ⚠ POSSIBLE DUPLICATE — проверить!").strip(" |")
        c.fill = ORANGE
        c.font = Font(name="Arial", size=8, color="FF843C0C", bold=True)


def apply_inv_update(ws, upd, wst=None):
    """
    Mark invoice as paid AND auto-create a Payment transaction if not already exists.
    Returns (found: bool, tx_created: bool, duplicate_row: int|None)
    """
    inv_no  = str(upd.get("invoice_no","")).strip().lower()
    status  = upd.get("new_status","✅ Paid")
    bg      = STAT_BG.get(status, YELLOW)

    for row in ws.iter_rows(min_row=5, max_col=11):
        if not inv_no:
            continue
        cell_inv = str(row[1].value or "").strip().lower()
        cell_ref = str(row[8].value or "").strip().lower()
        # Match: extracted inv_no is substring of Excel inv_no OR vice versa,
        # OR extracted inv_no matches the ref field (col I)
        matched = (inv_no in cell_inv or
                   (cell_inv and cell_inv in inv_no) or
                   (cell_ref and (inv_no in cell_ref or cell_ref in inv_no)))
        if not matched:
            continue

        # ── Update invoice status ─────────────────────────────────────────
        row[6].value = status; sc(row[6], bg=bg, bold=True, align="center")
        date_paid = upd.get("date_paid",""); row[7].value = date_paid; sc(row[7], bg=bg)
        ref = upd.get("ref","")
        if ref: row[8].value = ref; sc(row[8], bg=bg, sz=8)

        # Write beneficiary to col K if provided (only our companies, not agent's)
        benef_upd = upd.get("beneficiary")
        if benef_upd and not is_agent_company_str(benef_upd):
            row[10].value = benef_upd
            sc(row[10], bg=bg, sz=9)

        # ── Only auto-create transaction when marking as Paid ────────────
        if status != "✅ Paid" or wst is None:
            return True, False, None

        # Determine amount/ccy: SWIFT > invoice > fallback
        swift_amt  = upd.get("swift_amount")
        swift_ccy  = upd.get("swift_ccy")
        swift_date = upd.get("swift_date") or date_paid

        inv_ccy    = str(row[3].value or "")
        try: inv_amt = float(row[4].value or 0)
        except: inv_amt = 0.0

        if swift_amt and swift_ccy:
            tx_amt = float(swift_amt); tx_ccy = swift_ccy; tx_date = swift_date
            src = "SWIFT"
        elif inv_amt:
            tx_amt = inv_amt; tx_ccy = inv_ccy; tx_date = date_paid
            src = "инвойс"
        else:
            return True, False, None  # no amount at all — skip

        payee = str(row[2].value or "")

        # ── Dedup check ───────────────────────────────────────────────────
        dup_row = _find_duplicate_tx(wst, payee, tx_ccy, tx_amt, tx_date, ref=ref)
        if dup_row:
            # Transaction already exists — just add ref to its notes
            c = wst.cell(dup_row, 12)
            cur = str(c.value or "")
            if ref and ref not in cur:
                c.value = (cur + f" | ref: {ref}").strip(" |")
            log.info(f"Invoice {inv_no}: transaction already exists at row {dup_row}, skipping creation")
            return True, False, dup_row

        # ── Create transaction ────────────────────────────────────────────
        inv_no_display = str(row[1].value or "")
        tx = {
            "date":        tx_date,
            "type":        "Payment",
            "description": f"{inv_no_display} — {payee}",
            "payee":       payee,
            "ccy":         tx_ccy,
            "amount":      tx_amt,
            "fx_rate":     upd.get("swift_fx") or None,
            "comm":        None,
            "notes":       f"inv={inv_no_display} | Автозапись из инвойса ({src})" + (f" | ref: {ref}" if ref else ""),
        }
        # Inherit beneficiary: upd field has priority over invoice col K
        inv_benef = row[10].value if len(row) > 10 else None
        final_benef = (upd.get("beneficiary") or inv_benef)
        if final_benef and not is_agent_company_str(final_benef):
            tx["beneficiary"] = final_benef
        new_row = find_last_row(wst) + 1
        apply_tx_row(wst, new_row, tx)
        log.info(f"Invoice {inv_no}: auto-created transaction at row {new_row} ({src})")
        return True, True, None

    # ── Fallback: match by payee + amount when inv_no didn't match ──────────
    swift_amt = upd.get("swift_amount")
    swift_ccy = upd.get("swift_ccy")
    payee_hint = str(upd.get("payee") or "").strip().lower()
    if payee_hint and status == "✅ Paid":
        for row in ws.iter_rows(min_row=5, max_col=11):
            if not (row[0].value or row[1].value): continue
            if row[6].value == "✅ Paid": continue  # already paid
            cell_payee = str(row[2].value or "").strip().lower()
            if not cell_payee: continue
            payee_words = [w for w in payee_hint.split() if len(w) > 3]
            if not any(w in cell_payee for w in payee_words): continue
            # Amount match if swift_amount provided
            if swift_amt:
                try:
                    cell_amt = float(row[4].value or 0)
                    if abs(float(swift_amt) - cell_amt) / max(float(swift_amt), 1) > 0.01:
                        continue
                except: continue
            # Found by payee+amount fallback — update status
            bg2 = STAT_BG.get(status, YELLOW)
            row[6].value = status; sc(row[6], bg=bg2, bold=True, align="center")
            date_paid = upd.get("date_paid", ""); row[7].value = date_paid; sc(row[7], bg=bg2)
            ref = upd.get("ref", "")
            if ref: row[8].value = ref; sc(row[8], bg=bg2, sz=8)
            log.info(f"Invoice fallback match by payee '{payee_hint}' at row {row[0].row}")
            if wst is None: return True, False, None
            inv_ccy = str(row[3].value or "")
            try: inv_amt = float(row[4].value or 0)
            except: inv_amt = 0.0
            if swift_amt and swift_ccy:
                tx_amt = float(swift_amt); tx_ccy = swift_ccy; tx_date = upd.get("swift_date") or date_paid
                src = "SWIFT"
            elif inv_amt:
                tx_amt = inv_amt; tx_ccy = inv_ccy; tx_date = date_paid; src = "инвойс"
            else:
                return True, False, None
            payee_display = str(row[2].value or "")
            dup_row2 = _find_duplicate_tx(wst, payee_display, tx_ccy, tx_amt, tx_date, ref=ref)
            if dup_row2:
                return True, False, dup_row2
            inv_no_display = str(row[1].value or "")
            tx = {
                "date": tx_date, "type": "Payment",
                "description": f"{inv_no_display} — {payee_display}",
                "payee": payee_display, "ccy": tx_ccy, "amount": tx_amt,
                "fx_rate": upd.get("swift_fx") or None, "comm": None,
                "notes": f"Автозапись из инвойса ({src})" + (f" | ref: {ref}" if ref else ""),
            }
            # Inherit beneficiary: upd field has priority over invoice col K
            inv_benef2 = row[10].value if len(row) > 10 else None
            final_benef2 = (upd.get("beneficiary") or inv_benef2)
            if final_benef2 and not is_agent_company_str(final_benef2):
                tx["beneficiary"] = final_benef2
            new_row2 = find_last_row(wst) + 1
            apply_tx_row(wst, new_row2, tx)
            log.info(f"Invoice fallback: auto-created transaction at row {new_row2}")
            return True, True, None

    return False, False, None

def repair_invoice_f_column(wsi):
    """After row deletion, rewrite F column formulas/values with correct row references.
    Also recomputes USD values for any rows where F is still a formula."""
    for row in wsi.iter_rows(min_row=5, max_col=6):
        r = row[0].row
        if row[0].value is None and row[1].value is None: continue
        f_cell = row[5]
        ccy = row[3].value
        amt = row[4].value
        # Recompute as number if possible
        if isinstance(amt, (int,float)) and amt > 0 and ccy:
            usd = _compute_usd(wsi.parent, str(ccy), amt)
            if usd:
                f_cell.value = usd
                continue
        # Otherwise rewrite formula with correct row number
        if isinstance(f_cell.value, str) and f_cell.value.startswith('='):
            f_cell.value = (f'=IF(OR(E{r}="",E{r}="TBC"),"TBC",'
                           f'IFERROR(E{r}/VLOOKUP(D{r},Settings!$A$7:$B$25,2,FALSE()),E{r}))')


def add_new_invoice(ws, inv, last_row):
    r   = last_row + 1
    st  = inv.get("status","⏳ Pending")
    bg  = STAT_BG.get(st, YELLOW)
    ccy = inv.get("ccy","")
    amt = inv.get("amount")
    for col_i, val in enumerate([
        inv.get("date",""), inv.get("invoice_no",""), inv.get("payee",""),
        ccy, amt, None, st, inv.get("date_paid",""), inv.get("ref",""), inv.get("notes","")
    ], 1):
        c = ws.cell(r, col_i, val if val is not None else "")
        sc(c, bg=bg, wrap=(col_i in (3,10)), sz=9)
    # Store USD as computed number — survives openpyxl save (no formula cache issue)
    if isinstance(amt, (int,float)) and amt > 0 and ccy:
        usd = _compute_usd(ws.parent, ccy, amt)
        ws.cell(r, 6).value = usd if usd else 0
    elif str(amt or "").upper() == "TBC" or not amt:
        ws.cell(r, 6).value = "TBC"
    else:
        ws.cell(r, 6).value = (f'=IF(OR(E{r}="",E{r}="TBC"),"TBC",'
                               f'IFERROR(E{r}/VLOOKUP(D{r},Settings!$A$7:$B$25,2,FALSE()),E{r}))')
    ws.cell(r,6).number_format = '#,##0.00'; sc(ws.cell(r,6), bg=bg)
    # K: Beneficiary
    benef = inv.get("beneficiary") or None
    if benef is not None:
        sc(ws.cell(r, 11, benef), bg=bg, sz=9, wrap=False)
    ws.row_dimensions[r].height = 26


def _recalc_balance_chain(ws, from_row: int):
    """Recompute K (Balance) for all rows from from_row downward."""
    from openpyxl.styles import Font, PatternFill, Alignment
    YELLOW = PatternFill("solid", fgColor="FFFFF2CC")
    prev_bal = 0.0
    for pr in ws.iter_rows(min_row=5, max_row=from_row-1, max_col=11, values_only=True):
        if pr[10] is not None and isinstance(pr[10], (int, float)):
            prev_bal = float(pr[10])
    if prev_bal == 0.0:
        try:
            s = ws.parent["Settings"].cell(4, 3).value
            if s and isinstance(s, (int, float)): prev_bal = float(s)
        except: pass
    for r_cells in ws.iter_rows(min_row=from_row, max_col=11):
        r = r_cells[0].row
        if not ws.cell(r, 1).value: break
        net_val = ws.cell(r, 10).value
        try: net = float(net_val) if isinstance(net_val, (int, float)) else 0.0
        except: net = 0.0
        new_bal = round(prev_bal + net, 2)
        c = ws.cell(r, 11, new_bal)
        c.number_format = "#,##0.00"; c.fill = YELLOW
        c.font = Font(name="Arial", size=9, bold=True, color="1F3864")
        c.alignment = Alignment(horizontal="right", vertical="center")
        prev_bal = new_bal


def apply_transaction_update(ws, upd):
    """Update notes/status/fx_rate of existing transaction row by matching description."""
    match_desc = str(upd.get("match_description","")).lower()
    match_date = str(upd.get("match_date","")).strip()
    new_notes  = upd.get("new_notes","")
    confirmed  = upd.get("confirmed", False)
    new_fx     = upd.get("fx_rate")

    if not match_desc:
        return False

    for row in ws.iter_rows(min_row=5):
        r = row[0].row
        desc  = str(ws.cell(r,3).value or "").lower()
        date  = str(ws.cell(r,1).value or "")
        notes = str(ws.cell(r,12).value or "")

        keywords = [w for w in match_desc.split() if len(w) > 3]
        matches  = sum(1 for kw in keywords if kw in desc or kw in notes.lower())

        if matches >= 1:
            if match_date and match_date not in date:
                continue

            # FX rate correction with full recalculation
            if new_fx:
                try:
                    from openpyxl.styles import Font, PatternFill, Alignment
                    YELLOW = PatternFill("solid", fgColor="FFFFF2CC")
                    fx    = float(new_fx)
                    amt   = float(ws.cell(r, 6).value or 0)
                    tp    = str(ws.cell(r, 2).value or "")
                    comm  = float(ws.cell(r, 9).value or 0)
                    gross = round(amt / fx, 2) if fx else amt
                    net   = round(gross, 2) if tp in ("Deposit","Cash In") else round(-(gross/max(1-comm,0.0001)),2)

                    c = ws.cell(r, 7, round(fx, 5))
                    c.number_format="0.00000"; c.fill=YELLOW
                    c.font=Font(name="Arial",size=9,color="0000CC")
                    c.alignment=Alignment(horizontal="right",vertical="center")

                    c = ws.cell(r, 8, gross)
                    c.number_format="#,##0.00"; c.fill=YELLOW
                    c.font=Font(name="Arial",size=9)
                    c.alignment=Alignment(horizontal="right",vertical="center")

                    c = ws.cell(r, 10, round(net, 2))
                    c.number_format="#,##0.00"; c.fill=YELLOW
                    c.font=Font(name="Arial",size=9)
                    c.alignment=Alignment(horizontal="right",vertical="center")

                    _recalc_balance_chain(ws, r)
                    notes = notes.replace("⏳ ПРЕДВ. КУРС","").replace("⏳ PRELIMINARY RATE","").strip(" |")
                    notes += f" | Курс подтверждён агентом: {fx}"
                    log.info(f"FX rate updated row {r}: fx={fx}, gross={gross}, net={net}")
                except Exception as e:
                    log.error(f"FX rate update error: {e}")

            # Notes / confirmation update
            updated_notes = notes
            updated_notes = updated_notes.replace("⚠ UNCONFIRMED — ", "✅ CONFIRMED — ")
            updated_notes = updated_notes.replace("Agent did NOT confirm", "Agent CONFIRMED")
            updated_notes = updated_notes.replace("FOLLOW UP!", "")
            if new_notes:
                updated_notes = new_notes
            if confirmed:
                ts = datetime.now().strftime("%d.%m.%Y")
                if "CONFIRMED" not in updated_notes.upper():
                    updated_notes += f" | Подтверждено агентом {ts}"

            ws.cell(r,12).value = updated_notes.strip(" |")
            cur_desc = str(ws.cell(r,3).value or "")
            if "⚠ UNCONFIRMED" in cur_desc or "UNCONFIRMED" in cur_desc:
                ws.cell(r,3).value = cur_desc.replace("⚠ UNCONFIRMED — ","✅ ").replace("UNCONFIRMED","CONFIRMED")

            log.info(f"Transaction updated row {r}: {ws.cell(r,3).value}")
            return True

    log.warning(f"Transaction not found for update: {match_desc}")
    return False

def _check_all_duplicates(wst) -> list:
    """
    Scan all Transactions for possible duplicates.
    Returns list of (row_a, row_b, reason) tuples for flagging.
    """
    from datetime import timedelta
    rows = []
    for row_idx, row in enumerate(wst.iter_rows(min_row=5, max_col=12, values_only=True), start=5):
        if not row[0]: continue
        rows.append((row_idx, row))

    flags = []
    for i, (ra, a) in enumerate(rows):
        for rb, b in rows[i+1:]:
            if a[1] not in ("Payment","Deposit") or b[1] not in ("Payment","Deposit"): continue
            try:
                amt_a = float(a[5] or 0); amt_b = float(b[5] or 0)
            except: continue
            if a[4] != b[4] or amt_a == 0: continue
            if abs(amt_a - amt_b) / max(amt_a, 1) > 0.01: continue
            da = _parse_date(a[0]); db = _parse_date(b[0])
            if da and db and abs((da - db).days) > 10: continue
            pa = str(a[3] or a[2] or "").lower()
            pb = str(b[3] or b[2] or "").lower()
            words = [w for w in pa.split() if len(w) > 3]
            if not any(w in pb for w in words): continue
            flags.append((ra, rb, f"{a[4]} {amt_a:,.0f} | {a[3]} | {a[0]} vs {b[0]}"))
    return flags


def write_to_excel(data: dict):
    if not EXCEL_FILE.exists(): return 0,0,0,0,0,[]
    wb  = load_workbook(EXCEL_FILE)
    wst = wb["Transactions"]; wsi = wb["Invoices"]
    tx_a = inv_u = inv_a = 0
    tx_upd = 0
    auto_tx = 0
    dup_warnings = []

    for tu in data.get("transaction_updates", []):
        if apply_transaction_update(wst, tu): tx_upd += 1
    for tx in data.get("new_transactions", []):
        apply_tx_row(wst, find_last_row(wst) + 1, tx); tx_a += 1
    for upd in data.get("invoice_updates", []):
        found, tx_created, dup_row = apply_inv_update(wsi, upd, wst)
        if found:
            inv_u += 1
            if tx_created: auto_tx += 1
            if dup_row: dup_warnings.append(
                f"⚠ Транзакция для {upd.get('invoice_no','')} уже существует (строка {dup_row}) — не дублировал")
    for inv in data.get("new_invoices", []):
        add_new_invoice(wsi, inv, find_last_row(wsi)); inv_a += 1

    # Run duplicate scan across all transactions
    dup_pairs = _check_all_duplicates(wst)
    for ra, rb, reason in dup_pairs:
        _flag_duplicate(wst, ra, rb)
        dup_warnings.append(f"⚠ ДУБЛЬ: строки {ra} и {rb} — {reason}")

    # Recalc full balance chain after any tx additions to fix any gaps
    if tx_a > 0 or auto_tx > 0:
        first_new = find_last_row(wst) - (tx_a + auto_tx)
        if first_new >= 5:
            _recalc_balance_chain(wst, first_new)

    wb.save(EXCEL_FILE)
    return tx_a, inv_u, inv_a, tx_upd, auto_tx, dup_warnings

# ── Claude API ────────────────────────────────────────────────────────────────
async def ask_claude(prompt_or_content, system=None) -> str:
    """
    Send a request to Claude API.
    prompt_or_content: str (text-only) or list (multimodal content blocks).
    system: str (plain) or list (cacheable blocks from _build_parse_system_prompt).
    """
    if system is None:
        sys_payload = "You are a financial assistant. Respond in Russian."
    else:
        sys_payload = system  # str or list — API accepts both

    if isinstance(prompt_or_content, str):
        content = [{"type": "text", "text": prompt_or_content}]
    else:
        content = prompt_or_content  # already a list of content blocks

    # Add caching header only when system is a cacheable list
    headers = {
        "x-api-key": ANTHROPIC_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    if isinstance(sys_payload, list):
        headers["anthropic-beta"] = "prompt-caching-2024-07-31"

    async with httpx.AsyncClient(timeout=120) as client:
        r = await client.post(
            "https://api.anthropic.com/v1/messages",
            headers=headers,
            json={"model": "claude-opus-4-6", "max_tokens": 4000,
                  "temperature": 0,
                  "system": sys_payload,
                  "messages": [{"role": "user", "content": content}]},
        )
        data = r.json()
        if "error" in data:
            raise RuntimeError(f"Claude API error: {data['error']}")
        return data["content"][0]["text"]

def _build_parse_system_prompt() -> list:
    """
    Build the full system prompt for parse_messages / invoice parsing.
    Returns a cacheable list block for Claude API prompt caching.
    """
    context = load_context()
    excel_bal = get_balance_from_excel()
    bal_str = f"${excel_bal[0]:,.2f} (запись: {excel_bal[1]})" if excel_bal else "нет данных"
    unconfirmed = get_recent_unconfirmed()
    unconfirmed_str = "\n".join(unconfirmed) if unconfirmed else "нет"
    existing_inv = get_existing_invoices_list()

    text = f"""КОНТЕКСТ ПРОЕКТА (обязательно учитывай):
{context}

ТЕКУЩИЙ БАЛАНС В EXCEL: {bal_str}

НЕПОДТВЕРЖДЁННЫЕ ТРАНЗАКЦИИ (мы отправили, агент ещё не подтвердил):
{unconfirmed_str}

УЖЕ СУЩЕСТВУЮЩИЕ ИНВОЙСЫ В EXCEL (НЕ добавляй их снова!):
{existing_inv}

---
Из новых сообщений от финансового агента извлеки структурированные данные.

Верни ТОЛЬКО валидный JSON без markdown:
{{
  "new_transactions": [
    {{
      "date": "DD.MM.YYYY",
      "type": "Payment|Deposit|Cash Out|Cash In|❓ Unknown",
      "description": "краткое описание",
      "payee": "название получателя",
      "ccy": "AED|CNY|USD|EUR|SGD|RUB|INR",
      "amount": 12345.67,
      "fx_rate": null,
      "comm": null,
      "notes": "доп. инфо",
      "payer": null,
      "beneficiary": null
    }}
  ],
  "invoice_updates": [
    {{
      "invoice_no": "номер инвойса",
      "new_status": "✅ Paid|⏳ Pending|🔄 In Progress|⚠ Partial/Check|❓ Clarify",
      "date_paid": "DD.MM.YYYY",
      "ref": "референс SWIFT или платёжный",
      "swift_amount": null,
      "swift_ccy": null,
      "swift_date": null,
      "swift_fx": null,
      "beneficiary": null
    }}
  ],
  "new_invoices": [
    {{
      "date": "DD.MM.YYYY",
      "invoice_no": "номер",
      "payee": "получатель",
      "ccy": "USD",
      "amount": 12345.67,
      "status": "⏳ Pending",
      "notes": "",
      "beneficiary": null
    }}
  ],
  "transaction_updates": [
    {{
      "match_description": "ключевые слова из описания существующей транзакции",
      "match_date": "DD.MM.YYYY или пусто",
      "new_notes": "",
      "confirmed": true,
      "fx_rate": null
    }}
  ],
  "balance_reconciliation": {{
    "agent_stated_balance": null,
    "our_excel_balance": null,
    "difference": null,
    "difference_explained_by": [],
    "unexplained_difference": null
  }},
  "context_update": "краткая запись для контекста — что нового узнали из этих сообщений",
  "summary": "2-3 предложения — что нового произошло"
}}

Правила:
- Сообщение с балансом агента ("Остаток: X") — занеси в balance_reconciliation, не в транзакции
- "ИСПОЛНЕН", "received", "RCVD", "Поступление подтверждаем", "получили", "поступило" = подтверждение → invoice_updates, НЕ new_transactions
- Платёжка "in progress", "отправлено", "wire sent", "sent", "в обработке", "transfer initiated", "processing", "выслал", "awaiting confirmation", "initiating payment" →
  статус инвойса = "🔄 In Progress". Заполни ref/swift_amount/swift_ccy/swift_date если есть в платёжке.
  НЕ создавай транзакцию в new_transactions — пользователь выберет через кнопки.
- "исполнено", "executed", "completed", "SWIFT отправлен", "wire completed" → статус "✅ Paid". Транзакция создастся автоматически.
- Если агент подтверждает получение без деталей — ищи в контексте последнюю UNCONFIRMED/FOLLOW UP транзакцию и обновляй её статус на ✅ Paid
- SWIFT-детали оплаты: если в сообщении есть SWIFT/MT103/сумма перевода → заполни в invoice_updates:
  swift_amount = сумма из SWIFT (число)
  swift_ccy    = валюта из SWIFT
  swift_date   = дата валютирования из SWIFT (DD.MM.YYYY)
  swift_fx     = курс конвертации если указан в SWIFT (иначе null)
  ref          = референс SWIFT (например "PACS008...", "OR260224...")
  Приоритет суммы для записи транзакции: SWIFT > инвойс > текст сообщения
- НЕ добавляй транзакцию в new_transactions если инвойс помечается как оплаченный — транзакция создастся автоматически
- Депозиты от нас агенту = Deposit. Получатель депозита = конечный получатель денег, не агент и не BALKEMY
- BALKEMY, TROVECO, RAWRIMA, ASTENO = плательщики (наша сторона), а не получатели

PAYER (кто отправил, Transactions col M):
- Всегда одна из компаний агента (матчинг нечёткий, по подстроке):
  BALKEMY GENERAL TRADING (варианты: Balkemy, BALKEMY GT)
  TROVECO DMCC (варианты: Troveco)
  ELITESPHERE PTE LTD (варианты: Elitesphere)
  RAWRIMA FZCO (варианты: Rawrima)
  MASADAN TRADING (варианты: Masadan)
  GORNIK TRADING LTD (варианты: Gornik)
  NEXUS MARINE PTE LTD (варианты: Nexus Marine, Nexus — когда плательщик)
  ASTENO LOGISTICS FZCO (варианты: Asteno)
- "from Balkemy", "BALKEMY account", "Elitesphere→" — всё это payer = Elitesphere/Balkemy
- Cash Out / Cash In обычно null. Если не ясно → null

BENEFICIARY (для кого, Transactions col N / Invoices col K):
- ТОЛЬКО наши компании (распознавай нечётко по ключевым словам):
  MENA → любое название с "MENA" (MENA Terminals, MENA Marine и т.д.)
  TRADE X → варианты: TradeX, Trade-X, Trade X Middle East
  INCOMED → любое название с "Incomed"
  OIQLO / OILQO → варианты: Oiqlo, Oilqo, OIQLO Services
  MMI = Mercantile Maritime International → варианты: MMI, Mercantile Maritime Int
  MMR = Mercantile Maritime Resources    → варианты: MMR
  MMT = Mercantile Maritime Trading      → варианты: MMT
  MME = Mercantile Maritime Engineering  → варианты: MME
  Myanmar Petroleum Services → варианты: Myanmar Petroleum, Myanmar Petroleum Svcs
  Maritime Shipping → любое название с "Maritime Shipping"
  Asia Shipco → любое название с "Asia Shipco"
  Nexus → когда выступает бенефициаром (не плательщиком)
- Компании агента (Balkemy, Troveco, Elitesphere, Rawrima, Masadan, Gornik, Asteno)
  НИКОГДА не бенефициары → null
- Для инвойсов: внимательно читай на кого выписан / для кого услуга:
  "insurance for MENA vessel" → MENA
  "MMI annual filing" → MMI (= Mercantile Maritime International)
  "Balkemy NOC letter" → null (Balkemy — агент)
  "TradeX services" → TRADE X
  "TROVECO insurance" → null (Troveco — агент)
- Batch-платёж (один SWIFT на несколько инвойсов разным бенефициарам) →
  перечисли через запятую: "MMI, MMR, MMT" или "Myanmar Petroleum Svcs, MMI, MMR, MMT"
  НЕ писать "Multiple SG entities" или любые обобщения с географией
- Если не ясно → null (лучше пустое чем неверное)

ЧТЕНИЕ PDF ИНВОЙСОВ:
Если в content переданы документы (PDF) — читай их внимательно.
Из каждого инвойса извлекай:
- Invoice number (номер инвойса)
- Date (дата)
- Payee / Vendor (кому выписан, получатель денег)
- Amount + Currency (сумма и валюта)
- Beneficiary — для КОГО эта услуга/товар (не кто получает деньги, а конечный
  бенефициар: название судна, компании, проекта, человека)
  Примеры: "vessel MT CHEM RON" → MT CHEM RON
           "for account of MENA Terminals" → MENA TERMINALS
           "re: Trade X deal" → TRADE X
           "visa cancellation for John Doe" → John Doe
- Payer — кто из компаний агента платит (если указано в инвойсе или платёжке)

Если передан SWIFT / платёжное поручение — извлекай:
- Ordering customer (плательщик) → payer
- Beneficiary customer → beneficiary
- Amount, currency, date, reference

Приоритет данных: PDF документ > текст сообщения (caption/forwarded text)

- Кэш который агент нам доставляет = Cash Out
- Непонятное → ❓ Unknown

ЛОГИКА СВЕРКИ БАЛАНСА (если агент прислал остаток):
1. agent_stated_balance — сумма из сообщения агента в USD
2. our_excel_balance — последний баланс из Excel (из контекста)
3. difference = our_excel_balance - agent_stated_balance (положительное = мы считаем больше чем агент)
4. difference_explained_by — список транзакций из "НЕПОДТВЕРЖДЁННЫЕ ТРАНЗАКЦИИ" которые объясняют разницу.
   Пример: мы отправили $150k, агент ещё не подтвердил → это объясняет $150k разницы.
   Формат: ["$150,000 USD отправлено 24.02 — агент не подтвердил (Pacs.008)"]
5. unexplained_difference = difference минус сумма объяснённых транзакций
   Если unexplained_difference близко к 0 — всё сходится.
   Если большое — есть реальное расхождение которое надо уточнять у агента."""

    return [{"type": "text", "text": text, "cache_control": {"type": "ephemeral"}}]


async def parse_messages(msgs_text: str) -> dict:
    system = _build_parse_system_prompt()
    prompt = f"Новые сообщения для анализа:\n{msgs_text}"
    raw = await ask_claude(prompt, system=system)
    return json.loads(_clean_json(raw))


# ── Format confirmation message ───────────────────────────────────────────────
def format_confirmation(data: dict) -> str:
    """
    Format clean CFO-ready confirmation message.
    Two parts returned as (report, warnings) — use format_technical_warnings for second message.
    """
    lines = []

    # ── Header: date + delta ──────────────────────────────────────────────────
    rec = data.get("balance_reconciliation", {})
    agent_bal = rec.get("agent_stated_balance")
    excel_bal = rec.get("our_excel_balance")
    diff      = rec.get("difference")

    date_str = ""
    # Try to get date from first transaction or invoice update
    for tx in data.get("new_transactions", []):
        if tx.get("date"): date_str = tx["date"]; break
    if not date_str:
        for u in data.get("invoice_updates", []):
            if u.get("date_paid"): date_str = u["date_paid"]; break

    if date_str and diff is not None:
        delta_fmt = f"Δ ${abs(float(diff)):,.2f}" if isinstance(diff, (int, float)) else f"Δ {diff}"
        status = "✅" if isinstance(diff, (int, float)) and abs(float(diff)) < 100 else "⚠"
        lines.append(f"Сверка {date_str} · {delta_fmt} {status}")
    else:
        lines.append("Сверка ✅")

    # ── Balance block ─────────────────────────────────────────────────────────
    if agent_bal is not None:
        lines.append("")
        lines.append(f"Агент:  ${float(agent_bal):,.2f}")
        if excel_bal is not None:
            lines.append(f"Excel:  ${float(excel_bal):,.2f}")
        if diff is not None and isinstance(diff, (int, float)):
            sign = "+" if float(diff) >= 0 else ""
            lines.append(f"Δ:      {sign}${abs(float(diff)):,.2f}")

    # ── Incoming deposits ─────────────────────────────────────────────────────
    txs = data.get("new_transactions", [])
    if txs:
        lines.append(f"\n📥 Поступления ({len(txs)}):")
        for tx in txs:
            try:
                amt_fmt = f"{float(tx.get('amount', 0)):,.2f}"
            except (TypeError, ValueError):
                amt_fmt = str(tx.get('amount', '?'))
            ccy = tx.get('ccy', '')
            payee = tx.get('payee') or tx.get('payer') or ""
            payee_str = f" · {payee}" if payee and payee != "None" else ""
            lines.append(f"  {amt_fmt} {ccy}{payee_str}")

    # ── Paid invoices ─────────────────────────────────────────────────────────
    upds = [u for u in data.get("invoice_updates", []) if "Paid" in u.get("new_status", "")]
    prog = [u for u in data.get("invoice_updates", []) if "Paid" not in u.get("new_status", "")]

    if upds:
        lines.append(f"\n📤 Оплачено ({len(upds)} инвойс{'ов' if len(upds) != 1 else ''}):")
        for u in upds:
            payee = u.get("payee") or u.get("invoice_no", "?")
            try:
                amt = u.get("swift_amount") or u.get("amount")
                ccy = u.get("swift_ccy") or u.get("ccy", "")
                amt_fmt = f"{float(amt):,.2f} {ccy}" if amt else ""
            except (TypeError, ValueError):
                amt_fmt = str(amt or "")
            amt_str = f" · {amt_fmt}" if amt_fmt else ""
            # Add brief description if different from payee
            desc = str(u.get("description") or u.get("invoice_no") or "").strip()
            desc_str = f"  ← {desc}" if desc and desc.lower() != payee.lower() else ""
            lines.append(f"  {payee}{amt_str}{desc_str}")

    if prog:
        lines.append(f"\n🔄 Обновлено ({len(prog)}):")
        for u in prog:
            payee = u.get("payee") or u.get("invoice_no", "?")
            lines.append(f"  {payee} → {u.get('new_status', '')}")

    # ── New invoices ──────────────────────────────────────────────────────────
    invs = data.get("new_invoices", [])
    if invs:
        lines.append(f"\n🆕 Новые инвойсы ({len(invs)}):")
        for inv in invs:
            try:
                amt_fmt = f"{float(inv.get('amount', 0)):,.2f}"
            except (TypeError, ValueError):
                amt_fmt = "TBC"
            lines.append(f"  {inv.get('payee', '?')} · {amt_fmt} {inv.get('ccy', '')}")

    # (skipped duplicates not shown in CFO report — internal info only)

    return "\n".join(lines)


def format_technical_warnings(data: dict) -> str:
    """
    Format second message with technical warnings (only sent if non-empty).
    Shown only to the operator, not intended for CFO.
    """
    warnings = []

    # Invoices marked Paid in Excel but no transaction found
    for u in data.get("invoice_updates", []):
        if u.get("_warning"):
            inv_no = u.get("invoice_no", "?")
            payee  = u.get("payee") or inv_no
            warnings.append(f"  {payee} — Paid в Excel, транзакция не найдена")

    if not warnings:
        return ""

    lines = ["⚠ Требует проверки:"]
    lines.extend(warnings)
    return "\n".join(lines)

# ── Commands ──────────────────────────────────────────────────────────────────
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Привет! Я ассистент по платежам.\n\n"
        "Просто пиши мне — я понимаю обычный текст:\n"
        "  'какой баланс?'\n"
        "  'покажи pending инвойсы дороже $10k'\n"
        "  'добавь транзакцию: получили 10000 EUR наличными'\n"
        "  'исправь получателя в последней строке'\n\n"
        "Пересылай сообщения от агента → /update.\n\n"
        "Команды:\n"
        "/update  — обработать накопленные сообщения\n"
        "/edit    — изменить строку в Excel\n"
        "/delete  — удалить последнюю транзакцию\n"
        "/balance — текущий баланс\n"
        "/pending — неоплаченные инвойсы\n"
        "/unknown — неизвестные транзакции\n"
        "/summary — полный отчёт\n"
        "/excel   — скачать Excel\n"
        "/clear   — очистить очередь сообщений"
    )


def _build_confirmation_keyboard(data: dict, confirm_cb: str = "confirm_update") -> InlineKeyboardMarkup:
    """
    Build smart confirmation keyboard based on what's actually in data.
    Button label shows exactly what will be written.
    """
    txs  = data.get("new_transactions", [])
    upds = data.get("invoice_updates", [])
    invs = data.get("new_invoices", [])

    # chat_action always has something to confirm
    is_chat_action = data.get("type") == "chat_action"
    has_any = bool(txs or upds or invs or is_chat_action)
    rows = []

    if not has_any:
        rows.append([InlineKeyboardButton("❌ Закрыть", callback_data="cancel_update")])
        return InlineKeyboardMarkup(rows)

    # For chat_action use preview as button label
    if is_chat_action and not (txs or upds or invs):
        preview = data.get("preview", "выполнить")
        label = f"✅ {preview}"
        rows.append([InlineKeyboardButton(label, callback_data=confirm_cb)])
        rows.append([InlineKeyboardButton("✏️ Внести правку", callback_data="request_edit")])
        rows.append([InlineKeyboardButton("❌ Отмена", callback_data="cancel_update")])
        return InlineKeyboardMarkup(rows)

    # Build descriptive label
    parts = []
    if txs:
        parts.append(f"{len(txs)} тр.")
    if upds:
        paid_count = sum(1 for u in upds if "Paid" in u.get("new_status", ""))
        prog_count = len(upds) - paid_count
        if paid_count:
            parts.append(f"{paid_count} инв. → Paid")
        if prog_count:
            parts.append(f"{prog_count} инв. → обновить")
    if invs:
        parts.append(f"{len(invs)} нов. инв.")

    label = "✅ Записать: " + " + ".join(parts)
    rows.append([InlineKeyboardButton(label, callback_data=confirm_cb)])
    rows.append([InlineKeyboardButton("✏️ Внести правку", callback_data="request_edit")])
    rows.append([InlineKeyboardButton("❌ Отмена", callback_data="cancel_update")])
    return InlineKeyboardMarkup(rows)


async def apply_pending_edit(pending_data: dict, instruction: str) -> dict:
    """
    Apply user's text instruction as a patch to pending JSON.
    Returns patched data dict (without _awaiting_edit flag).
    """
    # Serialize pending without internal flags for Claude
    clean = {k: v for k, v in pending_data.items() if not k.startswith("_")}
    pending_str = json.dumps(clean, ensure_ascii=False, indent=2)

    prompt = f"""У тебя есть JSON с данными которые готовятся к записи в Excel.
Пользователь хочет внести правку перед записью.

ТЕКУЩИЙ JSON:
{pending_str}

ИНСТРУКЦИЯ ПОЛЬЗОВАТЕЛЯ:
{instruction}

Верни ПОЛНЫЙ исправленный JSON — точно такой же структуры, только с применёнными правками.
Правь только то что просит пользователь. Всё остальное оставь без изменений.
Верни ТОЛЬКО валидный JSON без markdown, без объяснений, без backticks."""

    raw = await ask_claude(prompt, system=(
        "You are a JSON patch assistant. "
        "Return ONLY the complete patched JSON, no markdown, no explanation."
    ))
    return json.loads(_clean_json(raw))


async def cmd_update(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    msgs = load_messages()
    if not msgs:
        await update.message.reply_text(
            "Нет накопленных сообщений. Перешли что-нибудь от агента.")
        return

    has_pdfs   = any(m.get("pdf_b64") for m in msgs)
    has_images = any(m.get("img_b64") for m in msgs)
    has_media  = has_pdfs or has_images
    if has_media:
        media_count = sum(1 for m in msgs if m.get("pdf_b64") or m.get("img_b64"))
        pdf_count   = sum(1 for m in msgs if m.get("pdf_b64"))
        img_count   = sum(1 for m in msgs if m.get("img_b64"))
        parts = []
        if pdf_count: parts.append(f"{pdf_count} PDF нативно")
        if img_count: parts.append(f"{img_count} фото")
        await update.message.reply_text(
            f"Анализирую {len(msgs)} сообщений (в т.ч. {', '.join(parts)})..."
        )
    else:
        await update.message.reply_text(f"Анализирую {len(msgs)} сообщений...")

    try:
        if has_media:
            content = _build_multimodal_content(msgs)
            system  = _build_parse_system_prompt()
            raw     = await ask_claude(content, system=system)
            data    = json.loads(_clean_json(raw))
        else:
            data = await parse_messages(_fmt(msgs))
    except Exception as e:
        await update.message.reply_text(f"Ошибка анализа: {e}")
        log.error(f"Parse error: {e}"); return

    txs  = data.get("new_transactions", [])
    upds = data.get("invoice_updates", [])
    invs = data.get("new_invoices", [])

    if not txs and not upds and not invs:
        await update.message.reply_text(
            f"Новых транзакций не найдено.\n\n{data.get('summary','')}")
        # Still update context
        if data.get("context_update"):
            update_context_after_update(data["context_update"])
        clear_messages()
        return

    # Dedup: remove transactions and invoice updates already in Excel
    data, skipped_txs = _dedup_transactions(data)
    data, skipped_inv = _dedup_invoice_updates(data)
    if skipped_txs:
        data["_skipped_txs"] = skipped_txs
    if skipped_inv:
        data["_skipped_inv_upds"] = skipped_inv

    # Re-check if anything remains after dedup
    txs  = data.get("new_transactions", [])
    upds = data.get("invoice_updates", [])
    invs = data.get("new_invoices", [])

    if not txs and not upds and not invs:
        rec = data.get("balance_reconciliation", {})
        agent_bal = rec.get("agent_stated_balance")
        excel_bal = rec.get("our_excel_balance")
        diff      = rec.get("difference")

        lines = []
        date_str = ""
        for item in (skipped_txs + skipped_inv):
            # extract date from skipped description if possible
            pass
        # Try to get date from skipped inv entries
        # Just use today as fallback
        if diff is not None and isinstance(diff, (int, float)):
            delta_fmt = f"Δ ${abs(float(diff)):,.2f}"
            status = "✅" if abs(float(diff)) < 100 else "⚠"
            lines.append(f"Сверка · {delta_fmt} {status}")
        else:
            lines.append("Сверка ✅")

        if agent_bal is not None:
            lines.append(f"\nАгент:  ${float(agent_bal):,.2f}")
            if excel_bal is not None:
                lines.append(f"Excel:  ${float(excel_bal):,.2f}")
            if diff is not None and isinstance(diff, (int, float)):
                sign = "+" if float(diff) >= 0 else ""
                lines.append(f"Δ:      {sign}${abs(float(diff)):,.2f}")

        lines.append("\n✅ Все операции уже в Excel")

        await update.message.reply_text("\n".join(lines))
        if data.get("context_update"):
            update_context_after_update(data["context_update"])
        clear_messages()
        return

    # Save pending and show confirmation
    save_pending(data)
    conf_text = format_confirmation(data)

    keyboard = _build_confirmation_keyboard(data)
    await update.message.reply_text(conf_text, reply_markup=keyboard)

    # Send technical warnings as a separate message (no keyboard)
    tech_warnings = format_technical_warnings(data)
    if tech_warnings:
        await update.message.reply_text(tech_warnings)


def apply_edit(data: dict) -> str:
    """Apply an edit command to Excel."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    sheet_name = data.get("sheet", "Transactions")
    action     = data.get("action", "update")
    row_n      = int(data.get("row_number", 0))
    changes    = data.get("changes", {})
    desc       = data.get("description", "")

    if not EXCEL_FILE.exists():
        return "Excel файл не найден."

    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]

    thin = Side(style="thin", color="BFBFBF")
    def B(): return Border(top=thin, bottom=thin, left=thin, right=thin)

    if action == "delete":
        ws.delete_rows(row_n)
        # After deletion, repair F column formulas/values so row refs stay correct
        if sheet_name == "Invoices":
            repair_invoice_f_column(ws)
        # For Transactions: recompute balance chain from deleted row onward
        if sheet_name == "Transactions":
            _recalc_balance_chain(ws, row_n)
        wb.save(EXCEL_FILE)
        return f"Строка {row_n} удалена.\n{desc}"

    # Map col names to indices
    col_map = {
        "col_A":1,"col_B":2,"col_C":3,"col_D":4,"col_E":5,
        "col_F":6,"col_G":7,"col_H":8,"col_I":9,"col_J":10,
        "col_K":11,"col_L":12,"col_M":13,"col_N":14
    }

    applied = []
    for col_name, val in changes.items():
        if val is None: continue
        col_idx = col_map.get(col_name)
        if not col_idx: continue
        cell = ws.cell(row_n, col_idx, val)
        cell.font = Font(name="Arial", size=9)
        cell.border = B()
        cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (3,12)))
        applied.append(f"{col_name}={val}")

    # After writing base columns, recalculate formula columns G/H/I/J/K
    # if this is a Transactions sheet row with amount and CCY set
    if sheet_name == "Transactions":
        tp  = ws.cell(row_n, 2).value or ""
        ccy = ws.cell(row_n, 5).value or "USD"
        try: amt = float(ws.cell(row_n, 6).value or 0)
        except: amt = 0.0

        if amt and tp:
            # FX rate: приоритет col_G из changes > Settings
            # Используем sentinel None — чтобы явный fx=1.0 не перетирался Settings
            fx = None
            if changes.get("col_G") is not None:
                try:
                    fx = float(changes["col_G"])
                except: pass
            if fx is None:
                fx = 1.0
                try:
                    for srow in wb["Settings"].iter_rows(min_row=7, max_row=25, values_only=True):
                        if srow[0] == ccy:
                            fx = float(srow[1]); break
                except: pass

            # Comm
            comm_map = {"Deposit":0.0,"Cash In":0.0,"Payment":0.005,"Cash Out":0.005,"❓ Unknown":0.005}
            comm = comm_map.get(tp, 0.005)
            try:
                i_val = ws.cell(row_n, 9).value
                if i_val not in (None, "", "0%"):
                    comm = float(str(i_val).replace("%","")) / (100 if "%" in str(i_val) else 1)
            except: pass

            gross = round(amt / fx, 2) if fx else amt
            net   = round(gross, 2) if tp in ("Deposit","Cash In") else round(-(gross/max(1-comm,0.0001)),2)

            # Previous balance
            prev_bal = 0.0
            for pr in ws.iter_rows(min_row=5, max_row=row_n-1, max_col=11, values_only=True):
                if pr[10] is not None and isinstance(pr[10], (int, float)):
                    prev_bal = float(pr[10])
            if prev_bal == 0.0:
                try:
                    s = wb["Settings"].cell(4, 3).value
                    if s and isinstance(s, (int,float)): prev_bal = float(s)
                except: pass

            YELLOW = PatternFill("solid", fgColor="FFFFF2CC")

            c = ws.cell(row_n, 7, round(fx, 5))
            c.number_format = "0.00000"; c.fill = YELLOW
            c.font = Font(name="Arial", size=9, color="0000CC")
            c.alignment = Alignment(horizontal="right", vertical="center")

            c = ws.cell(row_n, 8, gross)
            c.number_format = "#,##0.00"; c.fill = YELLOW
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="right", vertical="center")

            c = ws.cell(row_n, 9, comm)
            c.number_format = "0.0%"; c.fill = YELLOW
            c.font = Font(name="Arial", size=9, color="0000CC")
            c.alignment = Alignment(horizontal="right", vertical="center")

            # Write J as computed NUMBER — avoids formula reference bugs entirely
            c = ws.cell(row_n, 10, round(net, 2))
            c.number_format = "#,##0.00"; c.fill = YELLOW
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(horizontal="right", vertical="center")

            c = ws.cell(row_n, 11, round(prev_bal + net, 2))
            c.number_format = "#,##0.00"; c.fill = YELLOW
            c.font = Font(name="Arial", size=9, bold=True, color="1F3864")
            c.alignment = Alignment(horizontal="right", vertical="center")

            # Mark as preliminary rate if CCY is not USD (rate may differ from agent's actual)
            if ccy != "USD":
                notes_cell = ws.cell(row_n, 12)
                cur_notes = str(notes_cell.value or "")
                if "ПРЕДВ. КУРС" not in cur_notes and "PRELIMINARY" not in cur_notes:
                    sep = " | " if cur_notes else ""
                    notes_cell.value = cur_notes + sep + "⏳ ПРЕДВ. КУРС — уточнить у агента"

            # Propagate balance change to all subsequent rows
            _recalc_balance_chain(ws, row_n + 1)
            applied.append(f"G={round(fx,5)} H={gross} I={comm} J={round(net,2)} K={round(prev_bal+net,2)}")

    # Invoices: recompute F (USD equiv) when amount or CCY changed
    if sheet_name == "Invoices" and ("col_E" in changes or "col_D" in changes):
        ccy_inv = ws.cell(row_n, 4).value
        amt_inv = ws.cell(row_n, 5).value
        if isinstance(amt_inv, (int,float)) and amt_inv > 0 and ccy_inv:
            usd = _compute_usd(wb, str(ccy_inv), amt_inv)
            if usd:
                ws.cell(row_n, 6).value = usd
                ws.cell(row_n, 6).number_format = '#,##0.00'
                applied.append(f"F(USD)={usd}")

    wb.save(EXCEL_FILE)
    return f"Применено к строке {row_n}:\n" + "\n".join(f"  {a}" for a in applied) + f"\n\n{desc}"

async def callback_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "request_edit":
        data = load_pending()
        if not data:
            await query.edit_message_text("Нет данных для правки.")
            return
        data["_awaiting_edit"] = True
        save_pending(data)
        await query.edit_message_text(
            "✏️ Напиши что изменить.\n\n"
            "Примеры:\n"
            "— измени payee на ORIENT INSURANCE\n"
            "— beneficiary → MMI\n"
            "— дата 28.02.2026\n"
            "— сумма 17799.36 AED"
        )
        return

    if query.data == "cancel_update":
        clear_pending()
        await query.edit_message_text("Отменено. Сообщения не удалены — можешь /update снова.")
        return

    if query.data == "confirm_mark_paid_with_tx":
        # Legacy callback — redirect to confirm_update (Paid+tx logic is always applied now)
        query.data = "confirm_update"


    data = load_pending()
    if not data:
        await query.edit_message_text("Нет данных для записи.")
        return

    # Handle chat action
    if data.get("type") == "chat_action":
        action = data.get("action","")
        params = data.get("params",{})
        try:
            if action == "add_transaction":
                tx_data = {"new_transactions":[params],"invoice_updates":[],"new_invoices":[]}
                write_to_excel(tx_data)
                msg2 = f"Транзакция добавлена."
            elif action == "add_invoice":
                tx_data = {"new_transactions":[],"invoice_updates":[],"new_invoices":[params]}
                write_to_excel(tx_data)
                msg2 = f"Инвойс добавлен."
            elif action in ("edit_transaction","edit_invoice"):
                sheet = "Transactions" if action == "edit_transaction" else "Invoices"
                edit_data = {"type":"edit","sheet":sheet,
                             "action":"update",
                             "row_number":params.get("row_number"),
                             "changes":params.get("changes",{}),
                             "description":data.get("preview","")}
                msg2 = apply_edit(edit_data)
            elif action == "delete_transaction":
                edit_data = {"type":"edit","sheet":"Transactions",
                             "action":"delete",
                             "row_number":params.get("row_number"),
                             "changes":{},
                             "description":data.get("preview","")}
                msg2 = apply_edit(edit_data)
            elif action == "delete_invoice":
                edit_data = {"type":"edit","sheet":"Invoices",
                             "action":"delete",
                             "row_number":params.get("row_number"),
                             "changes":{},
                             "description":data.get("preview","")}
                msg2 = apply_edit(edit_data)
            elif action == "mark_invoice_paid":
                # Full invoice_update path — creates auto-transaction
                inv_data = {"new_transactions":[],"invoice_updates":[params],"new_invoices":[]}
                tx_a, inv_u, inv_a, tx_upd, auto_tx, dups = write_to_excel(inv_data)
                msg2 = f"Инвойс обновлён." + (f" Создана транзакция." if auto_tx else "")
            else:
                msg2 = f"Неизвестное действие: {action}"
        except Exception as e:
            await query.edit_message_text(f"Ошибка: {e}"); return

        clear_pending()
        await query.edit_message_text(msg2)
        if EXCEL_FILE.exists():
            await ctx.bot.send_document(
                chat_id=MY_CHAT_ID,
                document=EXCEL_FILE.open("rb"),
                filename=f"Agent_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                caption="Excel обновлён"
            )
        return

    # Handle /edit command
    if data.get("type") == "edit":
        try:
            result_text = apply_edit(data)
        except Exception as e:
            await query.edit_message_text(f"Ошибка редактирования: {e}")
            log.error(f"Edit error: {e}"); return
        clear_pending()
        await query.edit_message_text(result_text)
        if EXCEL_FILE.exists():
            await ctx.bot.send_document(
                chat_id=MY_CHAT_ID,
                document=EXCEL_FILE.open("rb"),
                filename=f"Agent_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                caption="Excel после редактирования"
            )
        return

    try:
        tx_a, inv_u, inv_a, tx_upd, auto_tx, dup_warnings = write_to_excel(data)
    except Exception as e:
        await query.edit_message_text(f"Ошибка записи в Excel: {e}")
        log.error(f"Excel write error: {e}"); return

    # Update context
    if data.get("context_update"):
        update_context_after_update(data["context_update"])

    clear_pending()
    clear_messages()

    result = (f"Excel обновлён!\n\n"
              f"Транзакций добавлено: {tx_a}"
              + (f" (+{auto_tx} авто из инвойсов)" if auto_tx else "") + "\n"
              f"Транзакций обновлено: {tx_upd}\n"
              f"Инвойсов обновлено: {inv_u}\n"
              f"Инвойсов добавлено: {inv_a}")
    if dup_warnings:
        result += "\n\n" + "\n".join(dup_warnings)
    await query.edit_message_text(result)

    if EXCEL_FILE.exists():
        await ctx.bot.send_document(
            chat_id=MY_CHAT_ID,
            document=EXCEL_FILE.open("rb"),
            filename=f"Agent_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            caption="Обновлённый Excel"
        )


async def cmd_edit(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    Universal Excel editing command. Examples:
      /edit получили 10,000 EUR наличными в Монако
      /edit поменяй получателя в последней строке на Dubai Insurance
      /edit статус Dubai Insurance — оплачен 25.02.2026
      /edit уточни валюту строки 42 — это BHD а не AED
      /edit удали последнюю строку
      /edit Шафранов оплачен $14,280 банковским переводом
    """
    text = " ".join(ctx.args).strip()
    if not text:
        await update.message.reply_text(
            "Примеры команд:\n\n"
            "ДОБАВИТЬ:\n"
            "/edit получили 10,000 EUR наличными в Монако\n"
            "/edit депозит $50,000 подтверждён агентом сегодня\n"
            "/edit Шафранов оплачен $14,280 банковским переводом\n\n"
            "ИЗМЕНИТЬ:\n"
            "/edit поменяй получателя Dubai Insurance на оплаченный статус\n"
            "/edit исправь валюту последней строки на AED сумма 19502\n"
            "/edit статус Шафранов — оплачен 25.02.2026\n\n"
            "УДАЛИТЬ:\n"
            "/edit удали последнюю транзакцию\n"
            "/edit удали последние 2 строки"
        )
        return

    # Check if it's a delete command
    text_lower = text.lower()
    if any(w in text_lower for w in ["удали", "удалить", "delete"]):
        n = 1
        for word in text_lower.split():
            try: n = int(word); break
            except: pass
        n = min(n, 5)
        if not EXCEL_FILE.exists():
            await update.message.reply_text("Excel файл не найден.")
            return
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Transactions"]
        deleted = []
        for _ in range(n):
            last = None
            for row in ws.iter_rows(min_row=5):
                if row[0].value is not None: last = row[0].row
            if last:
                desc = ws.cell(last,3).value or ""
                amt  = ws.cell(last,6).value or ""
                ccy  = ws.cell(last,5).value or ""
                deleted.append(f"{desc} | {amt} {ccy}")
                ws.delete_rows(last)
        wb.save(EXCEL_FILE)
        result = f"Удалено {len(deleted)} строк:\n" + "\n".join(f"- {d}" for d in deleted)
        await update.message.reply_text(result)
        if EXCEL_FILE.exists():
            await ctx.bot.send_document(chat_id=MY_CHAT_ID,
                document=EXCEL_FILE.open("rb"),
                filename="Agent_after_edit.xlsx", caption="Excel после правки")
        return

    # For all other edits — ask Claude to generate JSON
    context = load_context()

    # Get current last 5 transactions for context
    tx_context = ""
    if EXCEL_FILE.exists():
        try:
            wb2 = load_workbook(EXCEL_FILE, data_only=True)
            ws2 = wb2["Transactions"]
            rows = [r for r in ws2.iter_rows(min_row=5, values_only=True) if r[0] is not None]
            tx_context = "Последние транзакции в Excel:\n" + "\n".join(
                f"  [{r[0]}] {r[1]} | {r[3] or '?'} | {r[5]} {r[4]} | bal={r[10]}"
                for r in rows[-8:]
            )
            inv_rows = [r for r in wb2["Invoices"].iter_rows(min_row=5, values_only=True)
                       if r[0] is not None and r[6] and "Pending" in str(r[6])]
            if inv_rows:
                tx_context += "\n\nPending инвойсы:\n" + "\n".join(
                    f"  {r[1]} | {r[2]} | {r[4]} {r[3]} | {r[6]}"
                    for r in inv_rows[:10]
                )
        except Exception as e:
            log.error(f"Excel read for edit: {e}")

    prompt = f"""КОНТЕКСТ ПРОЕКТА:
{context}

{tx_context}

---
Пользователь хочет внести правку в Excel одной командой:
"{text}"

Это может быть:
1. Добавление новой транзакции
2. Изменение статуса инвойса (если упоминается получатель + "оплачен/paid/исполнен")
3. Исправление данных существующей записи (валюта, сумма, получатель)

Верни ТОЛЬКО валидный JSON без markdown:
{{
  "new_transactions": [
    {{
      "date": "DD.MM.YYYY",
      "type": "Payment|Deposit|Cash Out|Cash In|❓ Unknown",
      "description": "краткое описание",
      "payee": "получатель",
      "ccy": "AED|USD|EUR|CNY|SGD|RUB|INR|BHD",
      "amount": 0.0,
      "fx_rate": null,
      "comm": null,
      "notes": "добавлено вручную"
    }}
  ],
  "invoice_updates": [
    {{
      "invoice_no": "номер или название",
      "new_status": "✅ Paid|⏳ Pending|🔄 In Progress|⚠ Partial/Check|❓ Clarify",
      "date_paid": "DD.MM.YYYY",
      "ref": ""
    }}
  ],
  "new_invoices": [],
  "balance_reconciliation": {{}},
  "context_update": "краткая запись для контекста",
  "summary": "одна строка — что изменили"
}}

Правила:
- "получили кэш/наличные" = Cash Out (агент доставил нам)
- "оплатили/заплатили" = Payment
- "депозит/отправили агенту" = Deposit
- "оплачен/paid/исполнен" + название = invoice_updates
- BALKEMY/TROVECO/RAWRIMA = плательщики, не получатели
- Дата не указана → сегодня: {datetime.now().strftime("%d.%m.%Y")}
- Если правка касается существующей записи (не новая) → используй invoice_updates
- Если нужно добавить новую строку → new_transactions"""

    await update.message.reply_text("Анализирую...")

    try:
        raw = await ask_claude(prompt, system=(
            "You are a JSON extraction assistant. "
            "Return ONLY valid JSON, no markdown, no backticks."
        ))
        data = json.loads(_clean_json(raw))
    except Exception as e:
        await update.message.reply_text(f"Ошибка анализа: {e}")
        return

    conf_text = format_confirmation(data)
    keyboard = _build_confirmation_keyboard(data)
    save_pending(data)
    await update.message.reply_text(conf_text, reply_markup=keyboard)

# Keep /add as alias for /edit
async def cmd_add(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await cmd_edit(update, ctx)



async def handle_chat(update: Update, ctx: ContextTypes.DEFAULT_TYPE, user_text: str):
    """Handle free-form user messages as AI conversation."""
    context    = load_context()
    excel_sum  = get_excel_summary()
    history    = load_chat_history()
    existing_inv = get_existing_invoices_list()

    # Build messages for Claude
    recent_tx = get_recent_transactions_with_rows(10)

    system_prompt = f"""Ты финансовый ассистент для трекера платежей через агента.
Отвечай по-русски, кратко и по делу.

ТЕКУЩЕЕ СОСТОЯНИЕ EXCEL:
{excel_sum}

ПОСЛЕДНИЕ ТРАНЗАКЦИИ (с номерами строк для редактирования):
{recent_tx}

УЖЕ СУЩЕСТВУЮЩИЕ ИНВОЙСЫ (с номерами строк):
{existing_inv}

КОНТЕКСТ ПРОЕКТА:
{context}

Если пользователь просит ДЕЙСТВИЕ — ответь JSON:
{{
  "type": "action",
  "action": "add_transaction|add_invoice|edit_transaction|edit_invoice|delete_transaction|delete_invoice|mark_invoice_paid",
  "params": {{...}},
  "preview": "одна строка — что именно сделаем",
  "message": "текст ответа пользователю"
}}

Если просто ВОПРОС — ответь JSON:
{{
  "type": "text",
  "message": "твой ответ"
}}

ПАРАМЕТРЫ ПО ДЕЙСТВИЯМ:

add_transaction: date, type(Payment|Deposit|Cash Out|Cash In|❓ Unknown), description, payee, ccy, amount, fx_rate(null=из настроек), comm(null), notes, payer(опц.), beneficiary(опц.)
add_invoice: date, invoice_no, payee, ccy, amount, status(⏳ Pending), notes, beneficiary(опц.)
edit_transaction: row_number(из списка выше!), changes: {{col_X: value}}
  Колонки: col_A=Date, col_B=Type, col_C=Desc, col_D=Payee, col_E=CCY, col_F=Amt, col_G=FX, col_L=Notes, col_M=Payer, col_N=Beneficiary
edit_invoice: row_number(из списка инвойсов!), changes: {{col_X: value}}
  Колонки: col_A=Date, col_B=InvNo, col_C=Payee, col_D=CCY, col_E=Amt, col_G=Status, col_H=DatePaid, col_I=Ref, col_J=Notes, col_K=Beneficiary
delete_transaction: row_number
delete_invoice: row_number
mark_invoice_paid: invoice_no, new_status("✅ Paid"), date_paid, ref(опц.), swift_amount(опц.), swift_ccy(опц.) — ИСПОЛЬЗУЙ для отметки инвойса оплаченным, создаёт транзакцию автоматически

ВАЖНО: row_number ВСЕГДА берётся из списков выше, никогда не угадывай!"""

    messages = []
    # Add history
    for h in history:
        messages.append(h)
    # Add current message
    messages.append({"role": "user", "content": user_text})

    try:
        async with httpx.AsyncClient(timeout=60) as client:
            r = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": ANTHROPIC_KEY,
                         "anthropic-version": "2023-06-01",
                         "anthropic-beta": "prompt-caching-2024-07-31",
                         "content-type": "application/json"},
                json={"model": "claude-opus-4-6", "max_tokens": 1500,
                      "temperature": 0,
                      "system": [{"type": "text", "text": system_prompt,
                                  "cache_control": {"type": "ephemeral"}}],
                      "messages": messages},
            )
            raw = _clean_json(r.json()["content"][0]["text"])
            data = json.loads(raw)
    except Exception as e:
        log.error(f"Chat error: {e}")
        await update.message.reply_text(f"Ошибка: {e}")
        return

    msg_text = data.get("message", "")
    response_type = data.get("type", "text")

    if response_type == "action":
        action   = data.get("action","")
        params   = data.get("params", {})
        preview  = data.get("preview","")

        # Save pending action
        pending_data = {"type": "chat_action", "action": action,
                        "params": params, "preview": preview}
        save_pending(pending_data)

        keyboard = _build_confirmation_keyboard(pending_data)
        reply = f"{msg_text}\n\n📋 {preview}"
        await update.message.reply_text(reply, reply_markup=keyboard)
    else:
        await update.message.reply_text(msg_text)

    # Save to history
    history.append({"role": "user",      "content": user_text})
    history.append({"role": "assistant", "content": msg_text})
    save_chat_history(history)

async def cmd_delete(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Delete last N rows from Transactions sheet. Usage: /delete or /delete 2"""
    n = 1
    if ctx.args:
        try: n = int(ctx.args[0])
        except: pass
    n = min(n, 5)  # max 5 at once for safety

    if not EXCEL_FILE.exists():
        await update.message.reply_text("Excel файл не найден.")
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Transactions"]

    deleted = []
    for _ in range(n):
        last = None
        for row in ws.iter_rows(min_row=5):
            if row[0].value is not None:
                last = row[0].row
        if last:
            desc  = ws.cell(last, 3).value or ""
            date  = ws.cell(last, 1).value or ""
            amt   = ws.cell(last, 6).value or ""
            ccy   = ws.cell(last, 5).value or ""
            deleted.append(f"[{date}] {desc} | {amt} {ccy}")
            ws.delete_rows(last)

    wb.save(EXCEL_FILE)

    if deleted:
        msg_text = f"Удалено {len(deleted)} строк:\n" + "\n".join(f"- {d}" for d in deleted)
        await update.message.reply_text(msg_text)
        await ctx.bot.send_document(
            chat_id=MY_CHAT_ID,
            document=EXCEL_FILE.open("rb"),
            filename="Agent_after_delete.xlsx",
            caption="Excel после удаления"
        )
    else:
        await update.message.reply_text("Нет строк для удаления.")


async def cmd_edit(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """
    Natural language Excel editor. Examples:
      /edit поменяй получателя в последней строке на MAXIMUS WAY
      /edit строка с Port of Fujairah — смени статус на оплачен 25.02
      /edit удали последнюю транзакцию
      /edit Dubai Insurance — обновить сумму 197334.74 AED
      /edit $150k строка 41 — добавь ref OR26022400002178 в примечания
    """
    text = " ".join(ctx.args).strip()
    if not text:
        await update.message.reply_text(
            "Укажи что изменить. Примеры:\n\n"
            "/edit поменяй получателя в последней строке на MAXIMUS WAY\n"
            "/edit строка с Port of Fujairah — статус оплачен 25.02\n"
            "/edit удали последнюю транзакцию\n"
            "/edit Dubai Insurance — обновить сумму 197334 AED\n"
            "/edit строка 41 — добавь примечание: подтверждено агентом"
        )
        return

    if not EXCEL_FILE.exists():
        await update.message.reply_text("Excel файл не найден.")
        return

    # Read current Excel state to give Claude context
    wb_read = load_workbook(EXCEL_FILE, data_only=True)
    ws_read = wb_read["Transactions"]
    wi_read = wb_read["Invoices"]

    tx_rows = []
    for row in ws_read.iter_rows(min_row=5, values_only=True):
        if row[0] is not None:
            tx_rows.append(row)

    inv_rows = []
    for row in wi_read.iter_rows(min_row=5, values_only=True):
        if row[0] is not None or row[1] is not None:
            inv_rows.append(row)

    # Format for Claude
    tx_text = "\n".join(
        f"Row {i+5}: [{r[0]}] {r[1]} | {r[3] or '?'} | {r[5]} {r[4]} | bal={r[10]} | notes={r[11] or ''}" +
        (f" | payer={r[12]}" if len(r) > 12 and r[12] else "") +
        (f" | for={r[13]}" if len(r) > 13 and r[13] else "")
        for i, r in enumerate(tx_rows)
    )
    inv_text = "\n".join(
        f"Row {i+5}: [{r[0]}] inv={r[1]} | {r[2]} | {r[4]} {r[3]} | status={r[6]} | paid={r[7]}" +
        (f" | for={r[10]}" if len(r) > 10 and r[10] else "")
        for i, r in enumerate(inv_rows)
    )

    context = load_context()

    prompt = f"""КОНТЕКСТ:
{context}

ТЕКУЩИЕ ТРАНЗАКЦИИ (Transactions sheet, строки начиная с 5):
Колонки: A=Date, B=Type, C=Description, D=Payee, E=CCY, F=Amount, G=FX, H=GrossUSD, I=Comm%, J=NetUSD, K=Balance, L=Notes, M=Payer, N=Beneficiary
{tx_text}

ТЕКУЩИЕ ИНВОЙСЫ (Invoices sheet):
Колонки: A=Date, B=InvNo, C=Payee, D=CCY, E=Amount, F=USD, G=Status, H=DatePaid, I=Ref, J=Notes, K=Beneficiary
{inv_text}

КОМАНДА ПОЛЬЗОВАТЕЛЯ: {text}

Верни ТОЛЬКО валидный JSON без markdown:
{{
  "sheet": "Transactions|Invoices",
  "action": "update|delete",
  "row_number": 42,
  "changes": {{
    "col_A": null,
    "col_B": null,
    "col_C": null,
    "col_D": null,
    "col_E": null,
    "col_F": null,
    "col_G": null,
    "col_H": null,
    "col_I": null,
    "col_J": null,
    "col_K": null,
    "col_L": null
  }},
  "description": "одна строка — что именно меняем и почему"
}}

Правила:
- row_number: точный номер строки Excel (начиная с 5)
- changes: только те колонки которые нужно изменить, остальные null
- action=delete: удалить строку целиком
- Если команда непонятна или строка не найдена — верни {{"error": "описание проблемы"}}
- Не пересчитывай баланс — только меняй указанные поля

КОЛОНКИ Transactions: col_A=Date, col_B=Type, col_C=Description, col_D=Payee, col_E=CCY, col_F=Amount(число), col_G=FX, col_H=GrossUSD, col_I=Comm%, col_J=NetUSD, col_K=Balance, col_L=Notes, col_M=Payer, col_N=Beneficiary
КОЛОНКИ Invoices: col_A=Date, col_B=InvNo, col_C=Payee, col_D=CCY(валюта), col_E=Amount(ЧИСЛО!), col_F=USD_equiv, col_G=Status, col_H=DatePaid, col_I=Ref, col_J=Notes, col_K=Beneficiary

ВАЖНО для Invoices:
- col_D = валюта (AED/USD/EUR/etc) — СТРОКА
- col_E = сумма — ЧИСЛО (например 242022.05, не "AED"!)
- Никогда не пиши валюту в col_E — только число
- Payer/Beneficiary: наши юрлица (RAWRIMA FZCO, BALKEMY GENERAL TRADING, TROVECO DMCC, ELITESPHERE PTE LTD, NEXUS MARINE PTE LTD, GORNIK TRADING LTD и др.)"""

    await update.message.reply_text("Анализирую команду...")

    try:
        raw = await ask_claude(prompt, system=(
            "You are a JSON assistant. Return ONLY valid JSON, no markdown, no backticks."
        ))
        data = json.loads(_clean_json(raw))
    except Exception as e:
        await update.message.reply_text(f"Ошибка анализа: {e}")
        return

    if "error" in data:
        await update.message.reply_text(f"Не понял команду: {data['error']}")
        return

    # Show confirmation
    desc = data.get("description","")
    row_n = data.get("row_number")
    action = data.get("action","update")
    sheet = data.get("sheet","Transactions")
    changes = data.get("changes",{})

    non_null = {k:v for k,v in changes.items() if v is not None}
    changes_text = "\n".join(f"  {k}: {v}" for k,v in non_null.items()) if non_null else "удаление строки"

    confirm_text = (
        f"Команда: {desc}\n\n"
        f"Лист: {sheet}\n"
        f"Строка: {row_n}\n"
        f"Действие: {action}\n"
        f"Изменения:\n{changes_text}"
    )

    save_pending({"type": "edit", "sheet": sheet, "action": action,
                  "row_number": row_n, "changes": changes, "description": desc})

    keyboard = _build_confirmation_keyboard({"type": "edit"})
    await update.message.reply_text(confirm_text, reply_markup=keyboard)

async def cmd_balance(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    result = get_balance_from_excel()
    if result:
        bal, date = result
        await update.message.reply_text(
            f"БАЛАНС АГЕНТА (из Excel)\n${bal:,.2f} USD\nПоследняя запись: {date}")
    else:
        await update.message.reply_text("Баланс не найден. Попробуй /update")

async def cmd_pending(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    pending, usd_total, tbc_count = get_pending_invoices()
    tbc_note = f"\n(+ {tbc_count} инвойс(ов) с суммой TBC)" if tbc_count else ""
    text = (f"ОЖИДАЮТ ОПЛАТЫ ({len(pending)}):\n\n" +
            ("\n".join(pending) if pending else "нет") +
            (f"\n\nИТОГО: ~${usd_total:,.0f} USD{tbc_note}" if pending else ""))
    await update.message.reply_text(text)

async def cmd_unknown(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    unknowns = get_unknown_transactions()
    text = f"НЕИЗВЕСТНЫЕ ТРАНЗАКЦИИ ({len(unknowns)}):\n\n" + (
        "\n".join(unknowns) if unknowns else "нет — хорошо!")
    await update.message.reply_text(text)

async def cmd_context(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    context = load_context()
    if not context:
        await update.message.reply_text("Контекст пуст.")
        return
    # Telegram limit 4096 chars
    if len(context) > 3800:
        context = context[-3800:]
        await update.message.reply_text(
            f"(Показаны последние 3800 символов)\n\n{context}")
    else:
        await update.message.reply_text(context)

async def cmd_summary(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await _send_report(ctx.bot, triggered_manually=True)

async def cmd_excel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if EXCEL_FILE.exists():
        await ctx.bot.send_document(
            chat_id=MY_CHAT_ID,
            document=EXCEL_FILE.open("rb"),
            filename="Agent_Model.xlsx",
            caption="Актуальный Excel"
        )
    else:
        await update.message.reply_text("Excel файл не найден на сервере.")

async def cmd_clear(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    clear_messages()
    await update.message.reply_text("Накопленные сообщения очищены.")

# ── Message handler ───────────────────────────────────────────────────────────
async def handle_message(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    if not msg: return
    if msg.chat_id != MY_CHAT_ID:
        log.info(f"IGNORED chat_id={msg.chat_id} expected={MY_CHAT_ID}")
        return
    text     = msg.text or msg.caption or ""

    # ── Intercept _awaiting_edit BEFORE forwarded check ──────────────────────
    if text and not getattr(msg, "document", None):
        _pending_check = load_pending()
        if _pending_check.get("_awaiting_edit"):
            await update.message.reply_text("⏳ Применяю правку...")
            try:
                patched = await apply_pending_edit(_pending_check, text)
                patched.pop("_awaiting_edit", None)
            except Exception as e:
                await update.message.reply_text(f"Ошибка при правке: {e}")
                return
            save_pending(patched)
            conf_text = format_confirmation(patched)
            keyboard = _build_confirmation_keyboard(patched)
            await update.message.reply_text(conf_text, reply_markup=keyboard)
            return

    # If NOT a forwarded message and NOT a document — treat as chat
    is_forwarded = bool(
        getattr(msg, "forward_origin", None) or
        getattr(msg, "forward_sender_name", None) or
        getattr(msg, "forward_from", None)
    )
    if not is_forwarded and not msg.document and not msg.photo and text:
        await handle_chat(update, ctx, text)
        return

    # Handle both old and new telegram-bot API forward attributes
    sender = ""
    try:
        origin = msg.forward_origin
        if origin:
            if hasattr(origin, "sender_name"):
                sender = origin.sender_name or ""
            elif hasattr(origin, "sender_user"):
                u = origin.sender_user
                sender = (u.full_name or u.username or "") if u else ""
            elif hasattr(origin, "chat"):
                c = origin.chat
                sender = (c.title or c.username or "") if c else ""
    except Exception:
        pass
    if not sender:
        try:
            sender = msg.forward_sender_name or ""
        except Exception:
            sender = ""
    date_str = msg.date.strftime("%d.%m.%Y %H:%M")
    try:
        if msg.forward_date:
            date_str = msg.forward_date.strftime("%d.%m.%Y %H:%M")
    except Exception:
        pass
    file_n    = msg.document.file_name if msg.document else ""
    pdf_b64   = None
    pdf_text  = ""
    img_b64   = None
    img_media = None

    # Download photo (sent as photo, not file)
    if msg.photo:
        try:
            tg_file = await msg.photo[-1].get_file()  # largest size
            buf = io.BytesIO()
            await tg_file.download_to_memory(buf)
            img_b64   = base64.b64encode(buf.getvalue()).decode("utf-8")
            img_media = "image/jpeg"
            file_n    = file_n or "photo.jpg"
            log.info(f"Photo stored as b64: {len(buf.getvalue())//1024}KB")
        except Exception as e:
            log.error(f"Photo download error: {e}")

    # Download image sent as document (jpg/png/webp/gif)
    _IMG_EXTS = (".jpg", ".jpeg", ".png", ".webp", ".gif")
    if msg.document and file_n.lower().endswith(_IMG_EXTS):
        try:
            tg_file = await msg.document.get_file()
            buf = io.BytesIO()
            await tg_file.download_to_memory(buf)
            ext = file_n.lower().rsplit(".", 1)[-1]
            img_media = {"jpg": "image/jpeg", "jpeg": "image/jpeg",
                         "png": "image/png", "webp": "image/webp",
                         "gif": "image/gif"}.get(ext, "image/jpeg")
            img_b64   = base64.b64encode(buf.getvalue()).decode("utf-8")
            log.info(f"Image doc stored as b64: {file_n}, {len(buf.getvalue())//1024}KB")
        except Exception as e:
            log.error(f"Image download error: {e}")

    # Download PDF — store as base64 for native Claude API reading
    if msg.document and file_n.lower().endswith(".pdf"):
        try:
            tg_file = await msg.document.get_file()
            buf = io.BytesIO()
            await tg_file.download_to_memory(buf)
            raw_bytes = buf.getvalue()

            if len(raw_bytes) > 5 * 1024 * 1024:  # >5MB — warn + fallback
                await msg.reply_text(
                    f"⚠️ PDF {file_n} большой ({len(raw_bytes)//1024//1024}MB), "
                    f"использую текстовое извлечение"
                )
            else:
                pdf_b64 = base64.b64encode(raw_bytes).decode("utf-8")
                log.info(f"PDF stored as b64: {file_n}, {len(raw_bytes)//1024}KB")

            # pypdf as text fallback (always attempt for small files)
            if HAS_PDF and len(raw_bytes) < 10 * 1024 * 1024:
                try:
                    buf.seek(0)
                    reader = PdfReader(buf)
                    pages_text = [p.extract_text() or "" for p in reader.pages]
                    pdf_text = "\n".join(pages_text)[:3000]
                except Exception:
                    pass  # b64 native path will handle it

        except Exception as e:
            log.error(f"PDF download error: {e}")
            pdf_text = f"[PDF не удалось скачать: {e}]"

    entry = {"date": date_str, "sender": sender, "text": text, "file": file_n}
    if pdf_b64:
        entry["pdf_b64"] = pdf_b64
    if pdf_text:
        entry["pdf_content"] = pdf_text  # fallback only
    if img_b64:
        entry["img_b64"]   = img_b64
        entry["img_media"] = img_media
    save_message(entry)

    preview = text[:60] + ("…" if len(text) > 60 else "")
    parts   = [f"от {sender}"] if sender else []
    if file_n:
        if pdf_b64:
            pdf_note = " (PDF нативно ✓)"
        elif img_b64:
            pdf_note = " (фото нативно ✓)"
        elif pdf_text:
            pdf_note = " (текст извлечён)"
        else:
            pdf_note = ""
        parts.append(f"файл: {file_n}{pdf_note}")
    if preview: parts.append(f'"{preview}"')
    count = len(load_messages())
    await msg.reply_text(
        f"Сохранено | {' | '.join(parts)}\n"
        f"В очереди: {count} сообщений. Когда готов — /update"
    )

# ── Morning report ────────────────────────────────────────────────────────────
async def _send_report(bot: Bot, triggered_manually=False):
    today = datetime.now().strftime("%d.%m.%Y")
    msgs  = load_messages()
    updates_text = ""

    if msgs:
        try:
            data    = await parse_messages(_fmt(msgs))
            tx_a, inv_u, inv_a, tx_upd, auto_tx, dup_warnings = write_to_excel(data)
            if data.get("context_update"):
                update_context_after_update(data["context_update"])
            if tx_a + inv_u + inv_a > 0:
                updates_text = (f"Автообновлено: +{tx_a} транзакций, "
                                f"{inv_u} инвойсов обновлено, +{inv_a} новых\n\n")
            clear_messages()
        except Exception as e:
            log.error(f"Auto-update error: {e}")
            updates_text = f"(Ошибка автообновления: {e})\n\n"

    result  = get_balance_from_excel()
    bal_str = f"${result[0]:,.2f} USD (запись: {result[1]})" if result else "нет данных"
    pending, usd_total, tbc_count = get_pending_invoices()
    unknown = get_unknown_transactions()
    tbc_note = f"\n  (+ {tbc_count} инвойс(ов) с суммой TBC — не включены)" if tbc_count else ""
    pending_total_str = f"\nИТОГО К ОПЛАТЕ: ~${usd_total:,.0f} USD{tbc_note}"

    text = (f"ОТЧЁТ — {today}\n\n"
            f"{updates_text}"
            f"БАЛАНС: {bal_str}\n\n"
            f"ОЖИДАЮТ ОПЛАТЫ ({len(pending)}):\n"
            + ("\n".join(pending) if pending else "нет") +
            (pending_total_str if pending else "") +
            f"\n\nНЕИЗВЕСТНЫЕ ({len(unknown)}):\n"
            + ("\n".join(unknown) if unknown else "нет"))

    await bot.send_message(chat_id=MY_CHAT_ID, text=text)
    if EXCEL_FILE.exists():
        await bot.send_document(
            chat_id=MY_CHAT_ID,
            document=EXCEL_FILE.open("rb"),
            filename=f"Agent_{datetime.now().strftime('%Y%m%d')}.xlsx",
            caption="Актуальный Excel"
        )

async def morning_job(ctx: ContextTypes.DEFAULT_TYPE):
    await _send_report(ctx.bot)

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    Path("data").mkdir(exist_ok=True)
    app = Application.builder().token(BOT_TOKEN).post_init(post_init).build()
    for cmd, fn in [
        ("start", cmd_start), ("update", cmd_update),
        ("edit", cmd_edit),
        ("delete", cmd_delete),
        ("edit", cmd_edit),
        ("balance", cmd_balance), ("pending", cmd_pending),
        ("unknown", cmd_unknown), ("summary", cmd_summary),
        ("excel", cmd_excel), ("context", cmd_context), ("clear", cmd_clear)
    ]:
        app.add_handler(CommandHandler(cmd, fn))
    app.add_handler(CallbackQueryHandler(callback_confirm))
    app.add_handler(MessageHandler(filters.ALL & ~filters.COMMAND, handle_message))
    app.job_queue.run_daily(morning_job, time=time(hour=MORNING_HOUR, minute=0))
    log.info("Bot v3 started")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

async def post_init(app):
    """Called after bot is initialized — run startup checks."""
    _ensure_settings_usdt()

if __name__ == "__main__":
    main()
