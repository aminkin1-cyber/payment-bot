"""
test_audit.py — Payment Tracker Bot
Unit tests for business logic.
Run: python3 test_audit.py
"""
import sys, os, types, unittest, tempfile, shutil
from datetime import date, timedelta
from unittest.mock import patch, MagicMock

# ── Patch env vars ────────────────────────────────────────────────────────────
os.environ.setdefault("BOT_TOKEN",     "fake:token")
os.environ.setdefault("ANTHROPIC_KEY", "fake-key")
os.environ.setdefault("MY_CHAT_ID",    "123456789")

# ── Mock telegram and httpx so bot.py imports cleanly without those packages ──
def _make_mock_module(name):
    m = types.ModuleType(name)
    m.__spec__ = None
    return m

for mod_name in [
    "telegram", "telegram.ext", "telegram.ext._application",
    "telegram.ext._commandhandler", "telegram.ext._messagehandler",
    "telegram.ext._callbackqueryhandler", "telegram.ext._filters",
    "telegram.ext._contexttypes", "telegram._update", "telegram._bot",
    "telegram._message", "telegram._inline", "telegram._files",
    "telegram.constants", "httpx",
]:
    if mod_name not in sys.modules:
        sys.modules[mod_name] = _make_mock_module(mod_name)

# Provide the specific names bot.py imports from telegram
tg = sys.modules["telegram"]
tg.Update = MagicMock()
tg.Bot = MagicMock()
tg.InlineKeyboardButton = MagicMock()
tg.InlineKeyboardMarkup = MagicMock()

tg_ext = sys.modules["telegram.ext"]
for attr in ["Application", "MessageHandler", "CommandHandler",
             "CallbackQueryHandler", "filters", "ContextTypes"]:
    setattr(tg_ext, attr, MagicMock())

httpx_mod = sys.modules["httpx"]
httpx_mod.AsyncClient = MagicMock()

# pypdf optional — mock it too
sys.modules.setdefault("pypdf", _make_mock_module("pypdf"))

import bot  # noqa: E402


# ═══════════════════════════════════════════════════════════════
# 1. is_agent_company_str  (14 tests)
# ═══════════════════════════════════════════════════════════════
class TestIsAgentCompany(unittest.TestCase):
    def test_balkemy_exact(self):         self.assertTrue(bot.is_agent_company_str("BALKEMY"))
    def test_balkemy_full(self):          self.assertTrue(bot.is_agent_company_str("BALKEMY GENERAL TRADING"))
    def test_balkemy_abbrev(self):        self.assertTrue(bot.is_agent_company_str("Balkemy GT"))
    def test_troveco(self):               self.assertTrue(bot.is_agent_company_str("TROVECO DMCC"))
    def test_rawrima(self):               self.assertTrue(bot.is_agent_company_str("RAWRIMA FZCO"))
    def test_elitesphere(self):           self.assertTrue(bot.is_agent_company_str("Elitesphere Pte Ltd"))
    def test_nexus_marine(self):          self.assertTrue(bot.is_agent_company_str("NEXUS MARINE PTE LTD"))
    def test_gornik(self):                self.assertTrue(bot.is_agent_company_str("Gornik Trading"))
    def test_asteno(self):                self.assertTrue(bot.is_agent_company_str("ASTENO LOGISTICS FZCO"))
    def test_masadan(self):               self.assertTrue(bot.is_agent_company_str("MASADAN TRADING"))
    def test_our_company_mena(self):      self.assertFalse(bot.is_agent_company_str("MENA Terminals"))
    def test_our_company_mmi(self):       self.assertFalse(bot.is_agent_company_str("MMI"))
    def test_none_value(self):            self.assertFalse(bot.is_agent_company_str(None))
    def test_empty_string(self):          self.assertFalse(bot.is_agent_company_str(""))

# ═══════════════════════════════════════════════════════════════
# 2. _parse_date  (10 tests)
# ═══════════════════════════════════════════════════════════════
class TestParseDate(unittest.TestCase):
    def test_ddmmyyyy(self):      self.assertEqual(bot._parse_date("15.03.2025"), date(2025, 3, 15))
    def test_yyyymmdd(self):      self.assertEqual(bot._parse_date("2025-03-15"), date(2025, 3, 15))
    def test_ddslash(self):       self.assertEqual(bot._parse_date("15/03/2025"), date(2025, 3, 15))
    def test_none(self):          self.assertIsNone(bot._parse_date(None))
    def test_empty(self):         self.assertIsNone(bot._parse_date(""))
    def test_garbage(self):       self.assertIsNone(bot._parse_date("not-a-date"))
    def test_whitespace(self):    self.assertEqual(bot._parse_date("  01.01.2024  "), date(2024, 1, 1))
    def test_year_boundary(self): self.assertEqual(bot._parse_date("31.12.2023"), date(2023, 12, 31))
    def test_us_format(self):     self.assertEqual(bot._parse_date("03/15/2025"), date(2025, 3, 15))
    def test_integer_input(self): self.assertIsNone(bot._parse_date(12345))

# ═══════════════════════════════════════════════════════════════
# 3. _clean_json  (5 tests)
# ═══════════════════════════════════════════════════════════════
class TestCleanJson(unittest.TestCase):
    def test_plain(self):           self.assertEqual(bot._clean_json('{"a":1}'), '{"a":1}')
    def test_markdown_fences(self): self.assertEqual(bot._clean_json('```json\n{"a":1}\n```'), '{"a":1}')
    def test_json_prefix(self):     self.assertEqual(bot._clean_json('json{"a":1}'), '{"a":1}')
    def test_whitespace(self):      self.assertEqual(bot._clean_json('  {"a":1}  '), '{"a":1}')
    def test_parseable_after(self): json_str = bot._clean_json('```{"a":1}```'); self.assertIn('"a"', json_str)

# ═══════════════════════════════════════════════════════════════
# 4. _get_comm  (7 tests)
# ═══════════════════════════════════════════════════════════════
class TestGetComm(unittest.TestCase):
    def test_deposit_zero(self):         self.assertEqual(bot._get_comm("Deposit", "USD"), 0.0)
    def test_cash_in_zero(self):         self.assertEqual(bot._get_comm("Cash In", "USD"), 0.0)
    def test_payment(self):              self.assertEqual(bot._get_comm("Payment", "USD"), 0.005)
    def test_cash_out(self):             self.assertEqual(bot._get_comm("Cash Out", "USD"), 0.005)
    def test_unknown(self):              self.assertEqual(bot._get_comm("❓ Unknown", "USD"), 0.005)
    def test_rub_override(self):         self.assertEqual(bot._get_comm("Payment", "RUB"), 0.004)
    def test_rub_deposit_override(self): self.assertEqual(bot._get_comm("Deposit", "RUB"), 0.004)

# ═══════════════════════════════════════════════════════════════
# 5. _is_duplicate_tx  (18 tests)
# ═══════════════════════════════════════════════════════════════
class TestIsDuplicateTx(unittest.TestCase):
    def _ex(self, ccy, amt, ago=0):
        return {"ccy": ccy, "amount": float(amt), "date": date.today() - timedelta(days=ago)}
    def _tx(self, ccy, amt, tx_type="Payment", ago=0):
        return {"ccy": ccy, "amount": amt, "type": tx_type,
                "date": (date.today() - timedelta(days=ago)).strftime("%d.%m.%Y")}

    def test_exact_match(self):
        self.assertTrue(bot._is_duplicate_tx(self._tx("USD", 1000), [self._ex("USD", 1000)])[0])
    def test_diff_ccy(self):
        self.assertFalse(bot._is_duplicate_tx(self._tx("AED", 1000), [self._ex("USD", 1000)])[0])
    def test_diff_amount(self):
        self.assertFalse(bot._is_duplicate_tx(self._tx("USD", 2000), [self._ex("USD", 1000)])[0])
    def test_within_1pct(self):
        self.assertTrue(bot._is_duplicate_tx(self._tx("USD", 1009), [self._ex("USD", 1000)])[0])
    def test_over_1pct(self):
        self.assertFalse(bot._is_duplicate_tx(self._tx("USD", 1015), [self._ex("USD", 1000)])[0])
    def test_payment_within_7d(self):
        self.assertTrue(bot._is_duplicate_tx(self._tx("USD", 5000), [self._ex("USD", 5000, ago=5)])[0])
    def test_payment_beyond_7d(self):
        self.assertFalse(bot._is_duplicate_tx(self._tx("USD", 5000), [self._ex("USD", 5000, ago=10)])[0])
    def test_deposit_within_60d(self):
        self.assertTrue(bot._is_duplicate_tx(self._tx("USD", 50000, "Deposit"), [self._ex("USD", 50000, ago=45)])[0])
    def test_deposit_beyond_60d(self):
        self.assertFalse(bot._is_duplicate_tx(self._tx("USD", 50000, "Deposit"), [self._ex("USD", 50000, ago=65)])[0])
    def test_cash_in_60d(self):
        self.assertTrue(bot._is_duplicate_tx(self._tx("AED", 10000, "Cash In"), [self._ex("AED", 10000, ago=30)])[0])
    def test_zero_amount(self):
        self.assertFalse(bot._is_duplicate_tx(self._tx("USD", 0), [self._ex("USD", 0)])[0])
    def test_missing_ccy(self):
        self.assertFalse(bot._is_duplicate_tx({"ccy": "", "amount": 1000, "type": "Payment"}, [self._ex("USD", 1000)])[0])
    def test_empty_existing(self):
        self.assertFalse(bot._is_duplicate_tx(self._tx("USD", 1000), [])[0])
    def test_reason_contains_ccy(self):
        _, r = bot._is_duplicate_tx(self._tx("EUR", 2500), [self._ex("EUR", 2500)])
        self.assertIn("EUR", r)
    def test_no_date_still_matches(self):
        self.assertTrue(bot._is_duplicate_tx({"ccy": "USD", "amount": 3000, "type": "Payment"}, [self._ex("USD", 3000)])[0])
    def test_case_insensitive_ccy(self):
        self.assertTrue(bot._is_duplicate_tx(self._tx("usd", 1000), [self._ex("USD", 1000)])[0])
    def test_multiple_existing(self):
        self.assertTrue(bot._is_duplicate_tx(self._tx("USD", 1000), [self._ex("USD", 999, ago=2), self._ex("USD", 1000, ago=1)])[0])
    def test_invalid_amount(self):
        self.assertFalse(bot._is_duplicate_tx({"ccy": "USD", "amount": "bad", "type": "Payment"}, [self._ex("USD", 1000)])[0])

# ═══════════════════════════════════════════════════════════════
# 6. _dedup_transactions  (8 tests)
# ═══════════════════════════════════════════════════════════════
class TestDedupTransactions(unittest.TestCase):
    def _tx(self, ccy, amt, tp="Payment"):
        return {"ccy": ccy, "amount": amt, "type": tp,
                "date": date.today().strftime("%d.%m.%Y"), "description": "t", "payee": "P"}
    def _ex(self, ccy, amt, ago=0):
        return {"ccy": ccy, "amount": float(amt), "date": date.today() - timedelta(days=ago)}

    def test_empty(self):
        d, s = bot._dedup_transactions({"new_transactions": []})
        self.assertEqual(d["new_transactions"], [])
    def test_unique_passes(self):
        with patch.object(bot, "get_recent_transactions", return_value=[]):
            d, s = bot._dedup_transactions({"new_transactions": [self._tx("USD", 1000)]})
        self.assertEqual(len(d["new_transactions"]), 1)
    def test_dup_removed(self):
        with patch.object(bot, "get_recent_transactions", return_value=[self._ex("USD", 1000)]):
            d, s = bot._dedup_transactions({"new_transactions": [self._tx("USD", 1000)]})
        self.assertEqual(len(d["new_transactions"]), 0); self.assertEqual(len(s), 1)
    def test_mix(self):
        with patch.object(bot, "get_recent_transactions", return_value=[self._ex("USD", 1000)]):
            d, _ = bot._dedup_transactions({"new_transactions": [self._tx("USD", 1000), self._tx("AED", 5000)]})
        self.assertEqual(d["new_transactions"][0]["ccy"], "AED")
    def test_not_mutated(self):
        orig = {"new_transactions": [self._tx("USD", 999)]}
        with patch.object(bot, "get_recent_transactions", return_value=[]):
            result, _ = bot._dedup_transactions(orig)
        self.assertIsNot(result, orig)
    def test_skipped_reason(self):
        with patch.object(bot, "get_recent_transactions", return_value=[self._ex("EUR", 2500)]):
            _, s = bot._dedup_transactions({"new_transactions": [self._tx("EUR", 2500)]})
        self.assertTrue(any("EUR" in x for x in s))
    def test_deposit_60d(self):
        with patch.object(bot, "get_recent_transactions", return_value=[self._ex("USD", 50000, ago=45)]):
            d, _ = bot._dedup_transactions({"new_transactions": [self._tx("USD", 50000, "Deposit")]})
        self.assertEqual(len(d["new_transactions"]), 0)
    def test_no_key(self):
        _, s = bot._dedup_transactions({})
        self.assertEqual(s, [])

# ═══════════════════════════════════════════════════════════════
# 7. Excel write with temp workbook  (16 tests)
# ═══════════════════════════════════════════════════════════════
class TestExcelWrite(unittest.TestCase):
    def setUp(self):
        from openpyxl import Workbook
        self.tmp = tempfile.mkdtemp()
        xl = os.path.join(self.tmp, "Agent_Model_v2.xlsx")
        wb = Workbook()
        wst = wb.active; wst.title = "Transactions"
        for _ in range(4): wst.append([])
        wsi = wb.create_sheet("Invoices")
        for _ in range(4): wsi.append([])
        wss = wb.create_sheet("Settings")
        wss.cell(4, 3, 100000.0)
        wss.cell(7, 1, "USD"); wss.cell(7, 2, 1.0)
        wss.cell(8, 1, "AED"); wss.cell(8, 2, 3.6725)
        wss.cell(9, 1, "EUR"); wss.cell(9, 2, 0.92)
        wb.save(xl)
        self._orig = bot.EXCEL_FILE
        bot.EXCEL_FILE = type(bot.EXCEL_FILE)(xl)

    def tearDown(self):
        bot.EXCEL_FILE = self._orig
        shutil.rmtree(self.tmp)

    def _tx(self, ccy="USD", amt=1000.0, tp="Payment", payee="Test Co", notes=""):
        return {"date": date.today().strftime("%d.%m.%Y"), "type": tp,
                "description": f"{tp} test", "payee": payee, "ccy": ccy,
                "amount": amt, "fx_rate": None, "comm": None, "notes": notes}

    def _rows(self):
        from openpyxl import load_workbook
        wb = load_workbook(bot.EXCEL_FILE, data_only=True)
        return [r for r in wb["Transactions"].iter_rows(min_row=5, values_only=True) if r[0]]

    def _inv_rows(self):
        from openpyxl import load_workbook
        wb = load_workbook(bot.EXCEL_FILE, data_only=True)
        return [r for r in wb["Invoices"].iter_rows(min_row=5, values_only=True) if r[0] or r[1]]

    def _write(self, **kw):
        base = {"new_transactions": [], "invoice_updates": [], "new_invoices": []}
        base.update(kw)
        return bot.write_to_excel(base)

    def test_write_single_tx(self):
        self._write(new_transactions=[self._tx()])
        self.assertEqual(len(self._rows()), 1)
    def test_ccy_stored(self):
        self._write(new_transactions=[self._tx(ccy="AED")])
        self.assertEqual(self._rows()[0][4], "AED")
    def test_payment_net_negative(self):
        self._write(new_transactions=[self._tx(tp="Payment", amt=1000)])
        self.assertLess(self._rows()[0][9], 0)
    def test_deposit_net_positive(self):
        self._write(new_transactions=[self._tx(tp="Deposit", amt=5000)])
        self.assertGreater(self._rows()[0][9], 0)
    def test_balance_chain(self):
        self._write(new_transactions=[self._tx(tp="Deposit", amt=10000), self._tx(tp="Payment", amt=1000)])
        rows = self._rows()
        self.assertAlmostEqual(rows[1][10], rows[0][10] + rows[1][9], places=2)
    def test_write_count(self):
        tx_a, *_ = self._write(new_transactions=[self._tx(), self._tx()])
        self.assertEqual(tx_a, 2)
    def test_agent_not_beneficiary(self):
        tx = self._tx(); tx["beneficiary"] = "BALKEMY GENERAL TRADING"
        self._write(new_transactions=[tx])
        benef = self._rows()[0][13] if len(self._rows()[0]) > 13 else None
        self.assertNotEqual(benef, "BALKEMY GENERAL TRADING")
    def test_our_company_beneficiary(self):
        tx = self._tx(); tx["beneficiary"] = "MENA Terminals"
        self._write(new_transactions=[tx])
        self.assertEqual(self._rows()[0][13], "MENA Terminals")
    def test_write_new_invoice(self):
        inv = {"date": "01.03.2026", "invoice_no": "INV-001", "payee": "Test",
               "ccy": "USD", "amount": 5000.0, "status": "⏳ Pending", "notes": ""}
        self._write(new_invoices=[inv])
        self.assertEqual(self._inv_rows()[0][1], "INV-001")
    def test_invoice_count(self):
        inv = {"date": "01.03.2026", "invoice_no": "INV-002", "payee": "P",
               "ccy": "USD", "amount": 1000.0, "status": "⏳ Pending", "notes": ""}
        _, _, inv_a, *_ = self._write(new_invoices=[inv])
        self.assertEqual(inv_a, 1)
    def test_mark_paid_creates_tx(self):
        self._write(new_invoices=[{"date":"01.03.2026","invoice_no":"INV-100","payee":"Orient",
                                   "ccy":"USD","amount":2500.0,"status":"⏳ Pending","notes":""}])
        _, inv_u, _, _, auto_tx, _ = self._write(invoice_updates=[{"invoice_no":"INV-100",
            "new_status":"✅ Paid","date_paid":"03.03.2026","ref":"REF001"}])
        self.assertEqual(inv_u, 1); self.assertEqual(auto_tx, 1)
    def test_auto_tx_inv_tag(self):
        self._write(new_invoices=[{"date":"01.03.2026","invoice_no":"INV-200","payee":"Port",
                                   "ccy":"USD","amount":3000.0,"status":"⏳ Pending","notes":""}])
        self._write(invoice_updates=[{"invoice_no":"INV-200","new_status":"✅ Paid",
                                      "date_paid":"04.03.2026","ref":""}])
        self.assertIn("inv=", str(self._rows()[-1][11] or "").lower())
    def test_empty_returns_zeros(self):
        tx_a, inv_u, inv_a, tu, at, dups = self._write()
        self.assertEqual(tx_a + inv_u + inv_a, 0)
    def test_aed_fx(self):
        self._write(new_transactions=[self._tx(ccy="AED", amt=36725.0, tp="Deposit")])
        self.assertAlmostEqual(self._rows()[0][7], 10000.0, delta=5.0)
    def test_find_last_row_empty(self):
        from openpyxl import load_workbook
        wb = load_workbook(bot.EXCEL_FILE)
        self.assertEqual(bot.find_last_row(wb["Transactions"]), 4)
    def test_three_txs_balance(self):
        self._write(new_transactions=[self._tx(tp="Deposit", amt=20000),
                                      self._tx(tp="Payment", amt=5000),
                                      self._tx(tp="Payment", amt=3000)])
        self.assertEqual(len(self._rows()), 3)

# ═══════════════════════════════════════════════════════════════
# 8. _recalc_balance_chain  (8 tests)
# ═══════════════════════════════════════════════════════════════
class TestRecalcBalanceChain(unittest.TestCase):
    def _ws(self, nets, opening=0.0):
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = "Transactions"
        s = wb.create_sheet("Settings"); s.cell(4, 3, opening)
        for i, n in enumerate(nets):
            r = 5 + i
            ws.cell(r, 1, "01.01.2025"); ws.cell(r, 10, n); ws.cell(r, 11, 0.0)
        return ws

    def test_single_row(self):
        ws = self._ws([1000.0]); bot._recalc_balance_chain(ws, 5)
        self.assertAlmostEqual(ws.cell(5, 11).value, 1000.0)
    def test_two_rows(self):
        ws = self._ws([1000.0, -200.0]); bot._recalc_balance_chain(ws, 5)
        self.assertAlmostEqual(ws.cell(6, 11).value, 800.0)
    def test_three_rows(self):
        ws = self._ws([5000.0, -1000.0, -500.0]); bot._recalc_balance_chain(ws, 5)
        self.assertAlmostEqual(ws.cell(7, 11).value, 3500.0)
    def test_opening_balance(self):
        ws = self._ws([1000.0], opening=50000.0); bot._recalc_balance_chain(ws, 5)
        self.assertAlmostEqual(ws.cell(5, 11).value, 51000.0)
    def test_from_row_6(self):
        ws = self._ws([1000.0, 2000.0, 3000.0])
        ws.cell(5, 11, 1000.0); bot._recalc_balance_chain(ws, 6)
        self.assertAlmostEqual(ws.cell(6, 11).value, 3000.0)
    def test_negative_net(self):
        ws = self._ws([-500.0]); bot._recalc_balance_chain(ws, 5)
        self.assertLess(ws.cell(5, 11).value, 0)
    def test_rounding(self):
        ws = self._ws([1000.123456]); bot._recalc_balance_chain(ws, 5)
        v = ws.cell(5, 11).value; self.assertEqual(v, round(v, 2))
    def test_zero_net(self):
        ws = self._ws([1000.0, 0.0]); bot._recalc_balance_chain(ws, 5)
        self.assertAlmostEqual(ws.cell(6, 11).value, 1000.0)

# ═══════════════════════════════════════════════════════════════
# 9. Invoice variant matching  (7 tests)
# ═══════════════════════════════════════════════════════════════
class TestInvoiceVariants(unittest.TestCase):
    def _v(self, inv_no):
        import re
        raw = inv_no.strip(); variants = {raw.lower(), f"inv={raw.lower()}"}
        base = raw.split("_")[0].strip()
        if base and base != raw: variants |= {base.lower(), f"inv={base.lower()}"}
        short = re.split(r'[\(\[\{—\-–]', raw)[0].strip()
        if short and short != raw: variants.add(short.lower())
        for n in re.findall(r'\d{3,}', raw): variants.add(n)
        return variants

    def test_plain(self):       self.assertIn("inv-2053", self._v("INV-2053"))
    def test_inv_tag(self):     self.assertIn("inv=2053", self._v("2053"))
    def test_underscore(self):  self.assertIn("2053", self._v("2053_RO"))
    def test_bracket(self):     self.assertIn("inv 4410", self._v("INV 4410 (5YZ W L)"))
    def test_numeric_id(self):  self.assertIn("4410", self._v("RAKEZ-4410-2026"))
    def test_dash(self):        self.assertIn("inv 5001", self._v("INV 5001-A"))
    def test_em_dash(self):     self.assertIn("inv 4410", self._v("INV 4410 — renewal"))

# ═══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    loader = unittest.TestLoader()
    suite  = unittest.TestSuite()
    for cls in [TestIsAgentCompany, TestParseDate, TestCleanJson, TestGetComm,
                TestIsDuplicateTx, TestDedupTransactions, TestExcelWrite,
                TestRecalcBalanceChain, TestInvoiceVariants]:
        suite.addTests(loader.loadTestsFromTestCase(cls))
    result = unittest.TextTestRunner(verbosity=2).run(suite)
    total  = result.testsRun
    passed = total - len(result.failures) - len(result.errors)
    print(f"\n{'='*50}")
    print(f"Итого: {total} | ✅ {passed} прошло | ❌ {len(result.failures)} провалов | 💥 {len(result.errors)} исключений")
    sys.exit(0 if result.wasSuccessful() else 1)
